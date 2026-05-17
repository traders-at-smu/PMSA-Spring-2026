"""V6 CLI entry point — Kalshi × Polymarket cross-market arbitrage bot.

Usage
-----
    python main.py [--config config.json] COMMAND

Commands:
    validate   Check config file and pairs list, then exit.
    scan       Run one scan cycle, print results, then exit (no trades placed).
    run        Scan and trade continuously until Ctrl+C.

Configuration is loaded from config.json (or config.example.json as fallback).
Lines starting with // are treated as comments and ignored.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import os
from pathlib import Path

from colorama import Fore, Style, init as colorama_init

# Add src to path so we can import from it
sys.path.append(str(Path(__file__).parent / "src"))

from src.connectors import KalshiConnector, PolymarketConnector, load_pairs, get_latest_pairs_file
from src.fees import parse_formula, validate_formula
from src.bot import run_scan, run_loop, _resolve_poly_fee_rate
from src.scheduler import start_daily_scheduler

_SCRIPT_DIR = Path(__file__).parent
_SRC_DIR = _SCRIPT_DIR / "src"


# ── Config loading ────────────────────────────────────────────────────────────

def _strip_comments(text: str) -> str:
    """Remove lines whose first non-whitespace characters are //."""
    return "\n".join(
        line for line in text.splitlines()
        if not line.strip().startswith("//")
    )


def _deep_merge(base: dict, override: dict) -> dict:
    """Recursively merge override into base. override wins on scalar conflicts."""
    result = dict(base)
    for k, v in override.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = _deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def load_config(path: str) -> dict:
    example = _SCRIPT_DIR / "config.example.json"
    p = Path(path)

    # Start from config.example.json as the canonical defaults
    if example.exists():
        with example.open(encoding="utf-8") as f:
            cfg = json.loads(_strip_comments(f.read()))
    else:
        cfg = {}

    if p.exists() and p.resolve() != example.resolve():
        with p.open(encoding="utf-8") as f:
            user_cfg = json.loads(_strip_comments(f.read()))
            cfg = _deep_merge(cfg, user_cfg)

    # Load api-keys.json from the same directory as config, deep-merge on top
    keys_path = Path(path).parent / "api-keys.json"
    if not keys_path.is_absolute():
        keys_path = _SCRIPT_DIR / keys_path
    if keys_path.exists():
        with keys_path.open(encoding="utf-8") as f:
            keys = json.loads(_strip_comments(f.read()))
        cfg = _deep_merge(cfg, keys)

    # Ensure all data paths are prefixed with data/ if not already absolute
    data_keys = [
        "position_file", "entry_log", "exit_log", "opportunities_log",
        "expired_log", "failed_log", "bad_log"
    ]
    for key in data_keys:
        val = cfg.get(key)
        if val and not Path(val).is_absolute() and not val.startswith("data/"):
            cfg[key] = f"data/{val}"

    return cfg


def _validate_config(cfg: dict) -> list[str]:
    errors = []
    if not cfg.get("pairs_file"):
        errors.append("Missing 'pairs_file' in config")
    
    # If using automated CSV input, ensure the directory exists
    input_dir = cfg.get("input_files_dir", "input_files")
    if not Path(input_dir).exists():
        # Only a warning if we have a fallback pairs_file
        pass

    fees = cfg.get("fees", {})
    for venue in ("kalshi", "polymarket"):
        formula = fees.get(venue, {}).get("formula")
        if not formula:
            errors.append(f"Missing fee formula for {venue}")
        else:
            try:
                validate_formula(formula, venue)
            except ValueError as exc:
                errors.append(str(exc))
    return errors


def _compile_fee_fns(cfg: dict) -> None:
    """Replace string formulas in cfg with compiled callables."""
    fees = cfg.get("fees", {})
    for venue in ("kalshi", "polymarket"):
        formula = fees.get(venue, {}).get("formula")
        if formula:
            cfg["fees"][venue]["formula_fn"] = parse_formula(formula)


def _effective_mode(cfg: dict) -> str:
    mode = str(cfg.get("mode", "paper")).lower()
    if mode != "live":
        return mode
    k = cfg.get("kalshi", {})
    p = cfg.get("polymarket", {})
    creds_ok = bool(k.get("api_key") and k.get("private_key_base64") and p.get("private_key"))
    if not creds_ok:
        fallback = bool(cfg.get("fallback_to_paper", True))
        if not fallback:
            print(
                f"\n  {Fore.RED}FATAL: mode=live but credentials are missing "
                f"and fallback_to_paper=false.{Style.RESET_ALL}\n"
                f"  Set kalshi.api_key, kalshi.private_key_base64, polymarket.private_key\n"
                f"  Or set fallback_to_paper=true to allow automatic paper fallback.\n",
                file=sys.stderr,
            )
            sys.exit(1)
        print(f"  {Fore.YELLOW}! mode=live but credentials missing — falling back to paper{Style.RESET_ALL}")
        return "paper"
    return mode


def _validate_live_credentials(cfg: dict, kalshi, poly) -> bool:
    print("  [live] Validating credentials...")
    ok = True
    try:
        bal = kalshi.get_balance()
        print(f"  [live] {Fore.GREEN}✓{Style.RESET_ALL} Kalshi auth OK (balance=${bal['balance']:.2f})")
    except Exception as exc:
        print(f"  [live] {Fore.RED}✗ Kalshi auth FAILED: {exc}{Style.RESET_ALL}", file=sys.stderr)
        ok = False
    try:
        bal = poly.get_balance()
        print(f"  [live] {Fore.GREEN}✓{Style.RESET_ALL} Polymarket auth OK (cash=${bal['cash']:.2f})")
    except Exception as exc:
        print(f"  [live] {Fore.RED}✗ Polymarket auth FAILED: {exc}{Style.RESET_ALL}", file=sys.stderr)
        ok = False
    return ok


def _build_connectors(cfg: dict) -> tuple[KalshiConnector, PolymarketConnector]:
    k_cfg = cfg.get("kalshi", {})
    p_cfg = cfg.get("polymarket", {})
    proxy_config = cfg.get("proxy", {})
    kalshi = KalshiConnector(
        api_key=k_cfg.get("api_key", ""),
        private_key_base64=k_cfg.get("private_key_base64", ""),
        base_url=k_cfg.get("base_url", ""),
        proxy_config=proxy_config,
    )
    poly = PolymarketConnector(
        private_key=p_cfg.get("private_key", ""),
        api_key=p_cfg.get("api_key", ""),
        api_secret=p_cfg.get("api_secret", ""),
        api_passphrase=p_cfg.get("api_passphrase", ""),
        funder_address=p_cfg.get("funder_address", ""),
        clob_url=p_cfg.get("clob_url", ""),
        gamma_url=p_cfg.get("gamma_url", ""),
        proxy_config=proxy_config,
    )
    return kalshi, poly


def _health_check_proxy(cfg: dict) -> None:
    """Run once at startup to verify proxy connectivity and geo-routing."""
    import requests as _req
    proxy_cfg = cfg.get("proxy", {})
    if not proxy_cfg.get("enabled"):
        print(f"  {Style.DIM}[proxy] Proxy disabled — skipping health check{Style.RESET_ALL}")
        return
    h = proxy_cfg.get("host", "")
    port = proxy_cfg.get("port", 0)
    u = proxy_cfg.get("username", "")
    pw = proxy_cfg.get("password", "")
    verify = proxy_cfg.get("verify_ssl", True)
    proxy_url = f"http://{u}:{pw}@{h}:{port}"
    session = _req.Session()
    session.proxies.update({"http": proxy_url, "https": proxy_url})
    session.verify = verify

    # (a) Egress IP
    try:
        r = session.get("https://api.ipify.org", timeout=10)
        r.raise_for_status()
        print(f"  [proxy] {Fore.GREEN}✓{Style.RESET_ALL} Proxy egress IP: {r.text.strip()}")
    except Exception as exc:
        print(f"  [proxy] {Fore.YELLOW}⚠ Proxy egress IP check failed: {exc}{Style.RESET_ALL}")

    # (b) Polymarket geoblock
    try:
        r = session.get("https://polymarket.com/api/geoblock", timeout=10)
        r.raise_for_status()
        data = r.json()
        country = data.get("country", "?")
        blocked = data.get("blocked", "?")
        if blocked is True:
            print(f"  [proxy] {Fore.RED}✗ Polymarket geoblock: country={country} blocked={blocked}{Style.RESET_ALL}")
        else:
            print(f"  [proxy] {Fore.GREEN}✓{Style.RESET_ALL} Polymarket geoblock: country={country} blocked={blocked}")
    except Exception as exc:
        print(f"  [proxy] {Fore.YELLOW}⚠ Polymarket geoblock check failed: {exc}{Style.RESET_ALL}")

    # (c) Kalshi exchange status
    try:
        r = session.get("https://api.elections.kalshi.com/trade-api/v2/exchange/status", timeout=10)
        r.raise_for_status()
        data = r.json()
        exchange_active = data.get("exchange_active", "?")
        trading_active = data.get("trading_active", "?")
        print(f"  [proxy] {Fore.GREEN}✓{Style.RESET_ALL} Kalshi status: exchange_active={exchange_active} trading_active={trading_active}")
    except Exception as exc:
        print(f"  [proxy] {Fore.YELLOW}⚠ Kalshi status check failed: {exc}{Style.RESET_ALL}")


def _health_check_fee_rate(pairs: list[dict], poly: PolymarketConnector) -> bool:
    """Verify that the Polymarket fee-rate endpoint is reachable and returning data."""
    if not pairs:
        return True
    # Test with the first pair that has a valid token ID
    tid = None
    for test_pair in pairs:
        # Prefer pre-resolved token IDs from poly_token_ids (new CSV format)
        tids_list = test_pair.get("poly_token_ids") or []
        if tids_list and str(tids_list[0]).strip():
            candidate = str(tids_list[0]).strip()
            if len(candidate) > 10:  # real tokens are very long numeric strings
                tid = candidate
                break
        # Fall back to polymarket_yes_token_id (old Excel format)
        candidate = test_pair.get("polymarket_yes_token_id", "")
        if candidate and len(candidate) > 10:
            tid = candidate
            break
    if not tid:
        return True

    print(f"  [health] Checking Polymarket fee-rate connectivity...")
    try:
        rate = poly.get_fee_rate(tid)
        print(f"  [health] {Fore.GREEN}✓{Style.RESET_ALL} Fee-rate endpoint OK (rate={rate})")
        return True
    except Exception as exc:
        print(f"  [health] {Fore.RED}✗{Style.RESET_ALL} Fee-rate endpoint failed: {exc}")
        return False


def _print_banner(cfg: dict, pair_count: int) -> None:
    art_path = _SRC_DIR / "ascii-art(1).txt"
    if art_path.exists():
        with art_path.open(encoding="utf-8") as f:
            print(f"{Fore.CYAN}{f.read()}{Style.RESET_ALL}")

    mode = cfg.get("mode", "paper")
    color = Fore.GREEN if mode == "live" else Fore.YELLOW
    print(f"  {Style.BRIGHT}V6 Arbitrage Bot{Style.RESET_ALL}")
    print(f"  Pairs:     {pair_count}")
    print(f"  Mode:      {color}{mode.upper()}{Style.RESET_ALL}")
    
    if mode == "live":
        k_key = cfg.get("kalshi", {}).get("api_key", "")[:8]
        p_key = cfg.get("polymarket", {}).get("private_key", "")[:8]
        print(f"  Kalshi:    ...{k_key}")
        print(f"  Poly:      ...{p_key}")
    
    print(f"  Logs:      {cfg.get('entry_log')}")
    print(f"  Positions: {cfg.get('position_file')}")
    print("-" * 60)


def _merge_user_pairs(pairs_file: str, user_pairs_dirs: str | list | None) -> None:
    """Merge per-user staging CSVs into the master pairs Excel at startup."""
    if not user_pairs_dirs:
        return

    # Normalise to a list of resolved Paths
    raw = [user_pairs_dirs] if isinstance(user_pairs_dirs, str) else list(user_pairs_dirs)
    dirs: list[Path] = []
    pairs_parent = Path(pairs_file).parent
    for entry in raw:
        p = Path(entry)
        if not p.is_absolute():
            p = pairs_parent / entry
        if p.exists():
            dirs.append(p)

    if not dirs:
        return

    csv_files = sorted(f for d in dirs for f in d.glob("*/pairs.csv"))
    if not csv_files:
        return

    master = Path(pairs_file)
    # Merging only supports .xlsx master files
    if not master.exists() or not master.suffix.lower() == ".xlsx":
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [merge] openpyxl not installed — skipping user pairs merge", file=sys.stderr)
        return

    # Read the master Excel to build the existing-keys set
    def _load_master_keys(ws) -> set[tuple[str, str]]:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hmap = {str(h or "").strip().lower(): i + 1 for i, h in enumerate(headers)}
        poly_col = hmap.get("poly_market_id") or hmap.get("pair_id")
        kalshi_col = hmap.get("kalshi_market_id")
        keys: set[tuple[str, str]] = set()
        if poly_col is None or kalshi_col is None:
            return keys
        for r in range(2, ws.max_row + 1):
            pid = str(ws.cell(row=r, column=poly_col).value or "").strip()
            kid = str(ws.cell(row=r, column=kalshi_col).value or "").strip()
            if pid and kid:
                keys.add((pid, kid))
        return keys

    print(f"  [merge] Checking {len(csv_files)} user pairs directories...")
    wb = load_workbook(master)
    ws = wb.active
    existing = _load_master_keys(ws)
    added = 0

    for csv_path in csv_files:
        try:
            with csv_path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    pid = str(row.get("poly_market_id") or row.get("pair_id") or "").strip()
                    kid = str(row.get("kalshi_market_id") or "").strip()
                    if not pid or not kid or (pid, kid) in existing:
                        continue
                    
                    # Append new row to Excel
                    # This is a simplified append that assumes column order matches 
                    # Pairs_for_Kalshi_and_Polymarket.xlsx structure.
                    # In a production tool, you would map keys to exact column indices.
                    new_values = [row.get(h) for h in [
                        "poly_market_id", "kalshi_market_id", "title_clean", 
                        "category_tag", "poly_slug", "poly_url", "kalshi_url", "active"
                    ]]
                    ws.append(new_values)
                    existing.add((pid, kid))
                    added += 1
            
            # Clear the CSV after successful merge
            with csv_path.open("w", newline="", encoding="utf-8") as f:
                f.write("poly_market_id,kalshi_market_id,title_clean,category_tag,poly_slug,poly_url,kalshi_url,active\n")

        except Exception as exc:
            print(f"  [merge] Error processing {csv_path}: {exc}", file=sys.stderr)

    if added > 0:
        wb.save(master)
        print(f"  [merge] {Fore.GREEN}✓{Style.RESET_ALL} Added {added} new pair(s) to {master.name}")
    wb.close()


# ── Commands ──────────────────────────────────────────────────────────────────

def cmd_validate(args) -> int:
    cfg = load_config(args.config)
    print(f"  Validating config: {args.config}")
    errors = _validate_config(cfg)
    if errors:
        for e in errors:
            print(f"  {Fore.RED}✗{Style.RESET_ALL} {e}", file=sys.stderr)
        return 1
    
    # Try loading pairs
    latest = get_latest_pairs_file(cfg.get("input_files_dir", "input_files"))
    if latest:
        cfg["pairs_file"] = latest
        print(f"  [validate] Found latest CSV: {latest}")

    try:
        pairs = load_pairs(cfg["pairs_file"])
        print(f"  {Fore.GREEN}✓{Style.RESET_ALL} Config and pairs file ({len(pairs)} pairs) are valid.")
        return 0
    except Exception as exc:
        print(f"  {Fore.RED}✗{Style.RESET_ALL} Failed to load pairs: {exc}", file=sys.stderr)
        return 1


def cmd_scan(args) -> int:
    cfg = load_config(args.config)
    errors = _validate_config(cfg)
    if errors:
        for e in errors:
            print(f"  {Fore.RED}✗{Style.RESET_ALL} {e}", file=sys.stderr)
        return 1

    # Override pairs_file with the latest CSV from input_files if present
    latest = get_latest_pairs_file(cfg.get("input_files_dir", "input_files"))
    if latest:
        cfg["pairs_file"] = latest
        print(f"  [scan] Using latest CSV: {latest}")

    _compile_fee_fns(cfg)
    cfg["mode"] = _effective_mode(cfg)

    _merge_user_pairs(cfg["pairs_file"], cfg.get("user_pairs_dirs"))
    pairs = load_pairs(cfg["pairs_file"])
    if not pairs:
        print(f"{Fore.YELLOW}No active pairs found in '{cfg['pairs_file']}'{Style.RESET_ALL}")
        return 0

    _print_banner(cfg, len(pairs))
    kalshi, poly = _build_connectors(cfg)
    _health_check_proxy(cfg)
    if args.health_check and not _health_check_fee_rate(pairs, poly):
        print(f"  {Fore.RED}✗{Style.RESET_ALL} fee-rate health check failed — aborting scan")
        return 1

    # scan-only: no trades placed
    run_scan(pairs, kalshi, poly, cfg, execute=False, batch_start=1, total_pairs=len(pairs))
    return 0


def cmd_run(args) -> int:
    cfg = load_config(args.config)
    errors = _validate_config(cfg)
    if errors:
        for e in errors:
            print(f"  {Fore.RED}✗{Style.RESET_ALL} {e}", file=sys.stderr)
        return 1

    # Start the background 3 AM updater
    start_daily_scheduler(cfg)

    # Override pairs_file with the latest CSV from input_files if present
    latest = get_latest_pairs_file(cfg.get("input_files_dir", "input_files"))
    if latest:
        cfg["pairs_file"] = latest
        print(f"  [run] Using latest CSV: {latest}")

    _compile_fee_fns(cfg)
    cfg["mode"] = _effective_mode(cfg)

    _merge_user_pairs(cfg["pairs_file"], cfg.get("user_pairs_dirs"))
    pairs = load_pairs(cfg["pairs_file"])
    if not pairs:
        print(f"{Fore.YELLOW}No active pairs found in '{cfg['pairs_file']}'{Style.RESET_ALL}")
        return 0

    _print_banner(cfg, len(pairs))
    kalshi, poly = _build_connectors(cfg)
    if cfg.get("mode") == "live":
        if not _validate_live_credentials(cfg, kalshi, poly):
            print(
                f"\n  {Fore.RED}✗ Credential validation failed — bot will not start.{Style.RESET_ALL}\n"
                f"  Fix credentials in config.json and retry.\n",
                file=sys.stderr,
            )
            return 1
    _health_check_proxy(cfg)
    if args.health_check and not _health_check_fee_rate(pairs, poly):
        print(f"  {Fore.RED}✗{Style.RESET_ALL} fee-rate health check failed — aborting run")
        return 1

    run_loop(pairs, kalshi, poly, cfg)
    return 0


def cmd_balances(args) -> int:
    cfg = load_config(args.config)
    kalshi, poly = _build_connectors(cfg)
    
    print(f"\n{Style.BRIGHT}Account Balances{Style.RESET_ALL}")
    print("-" * 30)
    
    # Kalshi
    try:
        k_bal = kalshi.get_balance()
        print(f"Kalshi:")
        print(f"  Available Cash: ${Fore.GREEN}{k_bal['balance']:,.2f}{Style.RESET_ALL}")
        print(f"  Portfolio Val:  ${k_bal['portfolio_value']:,.2f}")
    except Exception as exc:
        print(f"Kalshi: {Fore.RED}Error fetching balance: {exc}{Style.RESET_ALL}")

    print("")

    # Polymarket
    try:
        p_bal = poly.get_balance()
        print(f"Polymarket:")
        print(f"  Available Cash: ${Fore.GREEN}{p_bal['cash']:,.2f}{Style.RESET_ALL}")
        print(f"  Total Balance:  ${p_bal['balance']:,.2f}")
    except Exception as exc:
        print(f"Polymarket: {Fore.RED}Error fetching balance: {exc}{Style.RESET_ALL}")
    
    print("-" * 30)
    return 0


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> int:
    colorama_init(autoreset=False)

    parser = argparse.ArgumentParser(
        description="Kalshi × Polymarket cross-market arbitrage bot  (v6)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python main.py validate\n"
            "  python main.py scan\n"
            "  python main.py run\n"
            "  python main.py --config /path/to/config.json run\n"
        ),
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config file (default: config.json). "
             "Falls back to config.example.json if not found.",
    )

    sub = parser.add_subparsers(dest="command", required=True)

    s = sub.add_parser("validate", help="Validate config and pairs file, then exit")
    s.set_defaults(func=cmd_validate)

    s = sub.add_parser(
        "scan",
        help="Run one scan cycle and print opportunities (no trades placed)",
    )
    s.add_argument(
        "--health-check",
        action="store_true",
        help="Verify Polymarket fee-rate endpoint before scanning",
    )
    s.set_defaults(func=cmd_scan)

    s = sub.add_parser(
        "run",
        help="Scan and execute trades continuously until Ctrl+C",
    )
    s.add_argument(
        "--health-check",
        action="store_true",
        help="Verify Polymarket fee-rate endpoint before running",
    )
    s.set_defaults(func=cmd_run)

    s = sub.add_parser("balances", help="Fetch and display exchange balances")
    s.set_defaults(func=cmd_balances)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
