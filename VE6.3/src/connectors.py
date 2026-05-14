"""API connectors for Kalshi and Polymarket, plus the Excel/CSV pairs loader.

All network calls are synchronous (requests library).
No credentials are required for orderbook data — only for placing live orders.
"""

from __future__ import annotations

import base64
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import padding

_KALSHI_BASE = "https://external-api.kalshi.com/trade-api/v2"
_POLY_CLOB = "https://clob.polymarket.com"
_POLY_GAMMA = "https://gamma-api.polymarket.com"
_TIMEOUT = 10  # seconds

def _slugify(text: str) -> str:
    text = text.lower().strip()
    slug = []
    prev_dash = False
    for ch in text:
        is_alnum = "a" <= ch <= "z" or "0" <= ch <= "9"
        if is_alnum:
            slug.append(ch)
            prev_dash = False
        else:
            if not prev_dash:
                slug.append("-")
                prev_dash = True
    out = "".join(slug).strip("-")
    return out

# ── Kalshi ────────────────────────────────────────────────────────────────────

class KalshiConnector:
    """Minimal Kalshi client: fetch orderbook quotes and place limit orders."""

    def __init__(self, api_key: str = "", private_key_base64: str = "", base_url: str = "", proxy_config: dict | None = None):
        # Use the public production URL by default — no key required for read access
        self.base = (base_url or _KALSHI_BASE).rstrip("/")
        self.api_key = api_key.strip()
        self._private_key = None
        self._event_url_cache: dict[str, str] = {}
        if private_key_base64.strip():
            key_bytes = base64.b64decode(private_key_base64.strip())
            self._private_key = serialization.load_pem_private_key(key_bytes, password=None)
        proxy_enabled = bool(proxy_config and proxy_config.get("enabled"))
        if proxy_enabled:
            h = proxy_config["host"]
            port = proxy_config["port"]
            u = proxy_config["username"]
            pw = proxy_config["password"]
            proxy_url = f"http://{u}:{pw}@{h}:{port}"
            self._proxies: dict | None = {"http": proxy_url, "https": proxy_url}
            self._verify: bool | str = proxy_config.get("verify_ssl", True)
        else:
            self._proxies = None
            self._verify = True
        self._session = requests.Session()
        if self._proxies:
            self._session.proxies.update(self._proxies)
        self._session.verify = self._verify

    def _event_url(self, event_ticker: str) -> str:
        if not event_ticker:
            return ""
        cached = self._event_url_cache.get(event_ticker)
        if cached is not None:
            return cached
        et = event_ticker.lower()
        url = f"https://kalshi.com/markets/{et}"  # always-valid event page fallback
        try:
            r = self._session.get(f"{self.base}/events/{event_ticker}", timeout=_TIMEOUT)
            r.raise_for_status()
            event = r.json().get("event", {})
            title = str(event.get("title") or "").strip()
            if title:
                slug = _slugify(title)
                if slug:
                    url = f"https://kalshi.com/markets/{et}/{slug}"
        except Exception:
            pass  # keep the event-page fallback
        self._event_url_cache[event_ticker] = url
        return url

    def _signed_headers(self, method: str, path: str) -> dict[str, str]:
        if not self._private_key or not self.api_key:
            raise RuntimeError(
                "Kalshi live trading requires 'api_key' and 'private_key_base64' in config"
            )
        ts_ms = str(int(time.time() * 1000))
        base_path = urlparse(self.base).path.rstrip("/")
        sign_path = path.split("?")[0]
        msg = f"{ts_ms}{method.upper()}{base_path}{sign_path}".encode()
        sig = self._private_key.sign(
            msg,
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.DIGEST_LENGTH,
            ),
            hashes.SHA256(),
        )
        return {
            "KALSHI-ACCESS-KEY": self.api_key,
            "KALSHI-ACCESS-TIMESTAMP": ts_ms,
            "KALSHI-ACCESS-SIGNATURE": base64.b64encode(sig).decode(),
            "Content-Type": "application/json",
        }

    @staticmethod
    def _parse_level(level: Any) -> tuple[int, int]:
        """Return (price_cents, qty) from a raw orderbook level.

        Handles both integer-cents format (e.g. 45 → 45¢) and decimal format
        (e.g. 0.45 → 45¢).  Kalshi's API has been observed returning either;
        using int() directly on a decimal truncates 0.45 → 0, making every
        derived ask resolve to $1.00.
        """
        if isinstance(level, dict):
            price = level.get("price", level.get("yes_price", level.get("no_price", 0)))
            size = level.get("count", level.get("quantity", level.get("size", 0)))
        elif isinstance(level, (list, tuple)) and len(level) >= 2:
            price, size = level[0], level[1]
        else:
            return 0, 0
        try:
            raw = float(price)
            # Prices > 1.0 are already in cents (e.g. 45); prices ≤ 1.0 are
            # decimal probabilities (e.g. 0.45) and must be scaled up.
            cents = int(round(raw * 100)) if raw <= 1.0 else int(round(raw))
            return cents, int(float(size))
        except (TypeError, ValueError):
            return 0, 0

    @classmethod
    def _derive_asks(cls, opposite_bids: list[Any]) -> list[dict[str, Any]]:
        """Kalshi only exposes bids; asks are derived from the opposite side.

        A bid for NO at 65 cents implies an ask for YES at 35 cents.
        """
        by_price: dict[float, int] = {}
        for level in opposite_bids:
            cents, qty = cls._parse_level(level)
            if qty <= 0:
                continue
            ask = round((100 - cents) / 100.0, 6)
            if 0.0 <= ask <= 1.0:
                by_price[ask] = by_price.get(ask, 0) + qty
        return [{"price": p, "size": by_price[p]} for p in sorted(by_price)]

    @classmethod
    def _derive_bids(cls, bids: list[Any]) -> list[dict[str, Any]]:
        """Convert raw Kalshi bid levels (yes/no) to normalized bid levels."""
        by_price: dict[float, int] = {}
        for level in bids:
            cents, qty = cls._parse_level(level)
            if qty <= 0:
                continue
            bid = round(cents / 100.0, 6)
            if 0.0 <= bid <= 1.0:
                by_price[bid] = by_price.get(bid, 0) + qty
        # descending for selling (best bid first)
        return [{"price": p, "size": by_price[p]} for p in sorted(by_price, reverse=True)]

    def get_quotes(self, ticker: str) -> dict[str, Any]:
        """Fetch and normalise orderbook quotes for a Kalshi market ticker.

        Makes two parallel requests:
          - GET /markets/{ticker}          → scalar bid/ask prices (V1-consistent)
          - GET /markets/{ticker}/orderbook → depth levels for slippage walking

        Returns:
            yes_bid, yes_ask, no_bid, no_ask (all floats in [0,1])
            depth.buy_yes  — ask levels for buying YES  [{price, size}, ...]
            depth.buy_no   — ask levels for buying NO   [{price, size}, ...]
        """
        def _fetch_orderbook() -> dict[str, Any]:
            r = self._session.get(f"{self.base}/markets/{ticker}/orderbook", timeout=_TIMEOUT)
            r.raise_for_status()
            return r.json()

        def _fetch_market() -> dict[str, Any]:
            r = self._session.get(f"{self.base}/markets/{ticker}", timeout=_TIMEOUT)
            r.raise_for_status()
            return r.json()

        with ThreadPoolExecutor(max_workers=2) as pool:
            ob_fut  = pool.submit(_fetch_orderbook)
            mkt_fut = pool.submit(_fetch_market)
            ob_data  = ob_fut.result()
            mkt_data = mkt_fut.result()

        # Orderbook depth — used for slippage depth-walk only
        ob = ob_data.get("orderbook_fp", {})
        yes_bids: list[Any] = ob.get("yes_dollars", [])
        no_bids:  list[Any] = ob.get("no_dollars", [])

        def _scalar(val: Any) -> float:
            """Parse a scalar price that may be integer cents (>1) or decimal (≤1)."""
            try:
                v = float(val or 0)
                return max(0.0, min(1.0, v / 100.0 if v > 1.0 else v))
            except (TypeError, ValueError):
                return 0.0

        # Scalar bid/ask from the market summary endpoint — same as V1
        mkt = mkt_data.get("market", {})
        event_ticker = str(mkt.get("event_ticker") or "").strip()
        event_url = self._event_url(event_ticker) if event_ticker else ""
        # Append the market ticker to form a direct 3-segment URL:
        # /markets/{event_ticker}/{event_slug}/{market_ticker}
        # Only do this when _event_url resolved past the bare fallback (i.e. has a slug).
        if event_url and ticker and event_ticker:
            _et_base = f"https://kalshi.com/markets/{event_ticker.lower()}"
            if event_url != _et_base:
                event_url = f"{event_url}/{ticker.lower()}"
        yes_bid = _scalar(mkt.get("yes_bid_dollars"))
        yes_ask = _scalar(mkt.get("yes_ask_dollars"))
        no_bid  = _scalar(mkt.get("no_bid_dollars"))
        no_ask  = _scalar(mkt.get("no_ask_dollars"))

        # Fall back to orderbook-derived prices only if market endpoint returns all zeros
        if yes_bid == 0.0 and yes_ask == 0.0 and no_bid == 0.0 and no_ask == 0.0:
            def _best_bid_ob(bids: list[Any]) -> float:
                best = max((self._parse_level(b)[0] for b in bids), default=0)
                return max(0.0, min(1.0, best / 100.0))
            yes_bid_ob = _best_bid_ob(yes_bids)
            no_bid_ob  = _best_bid_ob(no_bids)
            yes_bid = yes_bid_ob
            yes_ask = max(0.0, 1.0 - no_bid_ob)
            no_bid  = no_bid_ob
            no_ask  = max(0.0, 1.0 - yes_bid_ob)

        kalshi_title = str(mkt.get("subtitle") or mkt.get("title") or "").strip()

        return {
            "yes_bid": yes_bid,
            "yes_ask": yes_ask,
            "no_bid":  no_bid,
            "no_ask":  no_ask,
            "event_ticker": event_ticker,
            "event_url": event_url,
            "kalshi_title": kalshi_title,
            "depth": {
                "buy_yes": self._derive_asks(no_bids),   # to buy YES, cross NO bids
                "buy_no":  self._derive_asks(yes_bids),  # to buy NO, cross YES bids
                "sell_yes": self._derive_bids(yes_bids), # to sell YES, take YES bids
                "sell_no":  self._derive_bids(no_bids),  # to sell NO, take NO bids
            },
        }

    def get_balance(self) -> dict[str, float]:
        """Return available cash balance and total portfolio value."""
        path = "/portfolio/balance"
        headers = self._signed_headers("GET", path)
        res = self._session.get(f"{self.base}{path}", headers=headers, timeout=_TIMEOUT)
        res.raise_for_status()
        data = res.json()
        
        # Kalshi returns balance/portfolio_value in cents
        balance = float(data.get("balance", 0.0)) / 100.0
        portfolio = float(data.get("portfolio_value", 0.0)) / 100.0
        
        return {
            "balance": balance,
            "portfolio_value": portfolio,
        }

    def place_order(
        self,
        ticker: str,
        side: str,
        contracts: int,
        price: float,
        client_order_id: str,
        action: str = "buy",
    ) -> dict[str, Any]:
        """Place a limit IOC order on Kalshi at the quoted price.
        Fills up to contracts at the given price; unfilled remainder cancelled automatically."""
        if action not in {"buy", "sell"}:
            raise ValueError("action must be 'buy' or 'sell'")
        path = "/portfolio/orders"
        price_cents = int(round(price * 100))
        payload: dict[str, Any] = {
            "ticker": ticker,
            "client_order_id": client_order_id,
            "type": "limit",
            "action": action,
            "side": side,
            "count": contracts,
            "time_in_force": "immediate_or_cancel",
        }
        if side == "yes":
            payload["yes_price"] = price_cents
        else:
            payload["no_price"] = price_cents
        headers = self._signed_headers("POST", path)
        res = self._session.post(
            f"{self.base}{path}", json=payload, headers=headers, timeout=_TIMEOUT
        )
        res.raise_for_status()
        return res.json()

    def cancel_order(self, order_id: str) -> None:
        """Cancel an open order by order_id. Silently ignores 404 (already filled/cancelled)."""
        path = f"/portfolio/orders/{order_id}"
        headers = self._signed_headers("DELETE", path)
        res = self._session.delete(f"{self.base}{path}", headers=headers, timeout=_TIMEOUT)
        if res.status_code != 404:
            res.raise_for_status()


# ── Polymarket ────────────────────────────────────────────────────────────────

class PolymarketConnector:
    """Minimal Polymarket CLOB client: fetch quotes and place GTC limit orders."""

    def __init__(
        self,
        private_key: str = "",
        api_key: str = "",
        api_secret: str = "",
        api_passphrase: str = "",
        funder_address: str = "",
        clob_url: str = "",
        gamma_url: str = "",
        proxy_config: dict | None = None,
    ):
        # Use the public production URLs by default — no key required for read access
        self.clob_host = (clob_url or _POLY_CLOB).rstrip("/")
        self.gamma_host = (gamma_url or _POLY_GAMMA).rstrip("/")
        self.private_key = private_key.strip()
        self.api_key = api_key.strip()
        self.api_secret = api_secret.strip()
        self.api_passphrase = api_passphrase.strip()
        self.funder_address = funder_address.strip()
        self._client = None  # lazy-init on first live order
        proxy_enabled = bool(proxy_config and proxy_config.get("enabled"))
        if proxy_enabled:
            h = proxy_config["host"]
            port = proxy_config["port"]
            u = proxy_config["username"]
            pw = proxy_config["password"]
            proxy_url = f"http://{u}:{pw}@{h}:{port}"
            self._proxies: dict | None = {"http": proxy_url, "https": proxy_url}
            self._verify: bool | str = proxy_config.get("verify_ssl", True)
            # py_clob_client_v2.ClobClient does not expose a proxies kwarg.
            # Its internal requests calls honor HTTPS_PROXY/HTTP_PROXY env vars.
            os.environ["HTTPS_PROXY"] = proxy_url
            os.environ["HTTP_PROXY"] = proxy_url
        else:
            self._proxies = None
            self._verify = True
        self._session = requests.Session()
        if self._proxies:
            self._session.proxies.update(self._proxies)
        self._session.verify = self._verify

    def _ensure_client(self) -> None:
        if self._client is not None:
            return
        try:
            from py_clob_client_v2 import ClobClient, ApiCreds as V2ApiCreds, SignatureTypeV2  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "py-clob-client-v2 is required for live Polymarket trading. "
                "Install with: pip install py_clob_client_v2"
            ) from exc
        if not self.private_key:
            raise RuntimeError(
                "polymarket.private_key required for live mode"
            )
        creds = None
        if self.api_key and self.api_secret and self.api_passphrase:
            creds = V2ApiCreds(
                api_key=self.api_key,
                api_secret=self.api_secret,
                api_passphrase=self.api_passphrase,
            )
        self._client = ClobClient(
            host=self.clob_host,
            key=self.private_key,
            chain_id=137,
            creds=creds,
            signature_type=SignatureTypeV2.POLY_1271 if self.funder_address else None,
            funder=self.funder_address or None,
        )

    def _resolve_tokens(self, market_slug: str) -> tuple[str, str]:
        """Look up YES and NO token IDs for a market slug via Gamma API.

        Raises on non-binary outcomes (3+ outcomes) to avoid accidental
        default-to-index mapping that can invert or duplicate legs.
        """
        market = self._fetch_market(market_slug)
        outcomes, token_ids = self._extract_outcomes_tokens(market, market_slug)
        outcomes_lc = [o.strip().lower() for o in outcomes]
        if "yes" not in outcomes_lc or "no" not in outcomes_lc:
            raise RuntimeError(
                f"Polymarket market '{market_slug}' is non-binary outcomes={outcomes!r}; "
                "use explicit outcome mapping"
            )
        yes_idx = outcomes_lc.index("yes")
        no_idx = outcomes_lc.index("no")
        yes_id = str(token_ids[yes_idx]).strip()
        no_id = str(token_ids[no_idx]).strip()
        if not yes_id or not no_id or yes_id == no_id:
            raise RuntimeError(
                f"Polymarket market '{market_slug}' returned invalid token IDs: yes='{yes_id}', no='{no_id}'"
            )
        return yes_id, no_id

    def _fetch_market(self, market_slug: str) -> dict[str, Any]:
        """Fetch a Polymarket market payload for a given slug."""
        res = self._session.get(
            f"{self.gamma_host}/markets",
            params={"slug": market_slug},
            timeout=_TIMEOUT,
        )
        res.raise_for_status()
        payload = res.json()
        if isinstance(payload, list) and payload:
            market = payload[0]
        elif isinstance(payload, dict) and payload.get("conditionId"):
            market = payload
        else:
            raise RuntimeError(f"No Polymarket market found for slug '{market_slug}'")
        if not isinstance(market, dict):
            raise RuntimeError(f"Polymarket market payload malformed for slug '{market_slug}'")
        return market

    def get_balance(self) -> dict[str, float]:
        """Return available pUSD cash balance from Polymarket International CLOB (V2)."""
        self._ensure_client()
        try:
            from py_clob_client_v2 import BalanceAllowanceParams, AssetType  # type: ignore
            params = BalanceAllowanceParams(asset_type=AssetType.COLLATERAL)
            data = self._client.get_balance_allowance(params)
            if hasattr(data, "json"):
                data = data.json()
            cash = float(data.get("balance", 0.0)) / 1_000_000
            return {"cash": cash, "balance": cash}
        except Exception as exc:
            raise RuntimeError(f"Polymarket balance fetch failed: {exc}") from exc

    @staticmethod
    def _extract_outcomes_tokens(market: dict[str, Any], market_slug: str = "") -> tuple[list[str], list[str]]:
        raw_ids = market.get("clobTokenIds", [])
        if isinstance(raw_ids, str):
            raw_ids = json.loads(raw_ids)
        outcomes_raw = market.get("outcomes") or []
        if isinstance(outcomes_raw, str):
            try:
                outcomes_raw = json.loads(outcomes_raw)
            except json.JSONDecodeError:
                outcomes_raw = [outcomes_raw]
        outcomes = [str(o).strip() for o in outcomes_raw if str(o).strip()]
        token_ids = [str(t).strip() for t in (raw_ids or []) if str(t).strip()]
        if len(token_ids) < 2 or len(outcomes) < 2:
            label = f" '{market_slug}'" if market_slug else ""
            raise RuntimeError(
                f"Polymarket market{label} has insufficient outcomes/tokens "
                f"(outcomes={len(outcomes)}, tokens={len(token_ids)})"
            )
        return outcomes, token_ids

    def resolve_market_outcomes(self, market_slug: str) -> dict[str, Any]:
        """Return outcomes + token IDs (aligned) for a market slug."""
        market = self._fetch_market(market_slug)
        outcomes, token_ids = self._extract_outcomes_tokens(market, market_slug)
        return {
            "market": market,
            "outcomes": outcomes,
            "token_ids": token_ids,
        }

    def get_books(self, token_ids: list[str]) -> dict[str, dict[str, Any]]:
        """Fetch orderbooks for multiple Polymarket token IDs in parallel.

        Individual fetch failures are skipped with a warning rather than
        propagating — callers should check whether a token is present in the
        returned dict before using it.
        """
        token_ids = [str(t).strip() for t in token_ids if str(t).strip()]
        if not token_ids:
            return {}
        books: dict[str, dict[str, Any]] = {}
        with ThreadPoolExecutor(max_workers=min(10, len(token_ids))) as pool:
            futures = {pool.submit(self._fetch_book, tid): tid for tid in token_ids}
            for fut in futures:
                tid = futures[fut]
                try:
                    books[tid] = fut.result()
                except Exception as exc:
                    print(f"  [WARN] Polymarket book fetch failed for token …{tid[-8:]}: {exc} — retrying once")
                    try:
                        books[tid] = self._fetch_book(tid)
                    except Exception as exc2:
                        import requests as _req
                        if isinstance(exc2, _req.HTTPError) and getattr(exc2.response, "status_code", None) == 404:
                            raise  # re-raise 404 so run_scan logs it to expired_pairs
                        print(f"  [WARN] Polymarket book fetch failed again for token …{tid[-8:]}: {exc2}")
        return books

    @staticmethod
    def _parse_ask_levels(book: dict[str, Any]) -> list[dict[str, Any]]:
        by_price: dict[float, int] = {}
        for level in book.get("asks", []):
            try:
                if isinstance(level, dict):
                    p = float(level["price"])
                    q = int(float(level.get("size", level.get("quantity", 0))))
                elif isinstance(level, (list, tuple)):
                    p, q = float(level[0]), int(float(level[1]))
                else:
                    continue
            except (TypeError, ValueError, KeyError):
                continue
            if q > 0 and 0.0 <= p <= 1.0:
                rp = round(p, 6)
                by_price[rp] = by_price.get(rp, 0) + q
        return [{"price": p, "size": by_price[p]} for p in sorted(by_price)]

    @staticmethod
    def _parse_bid_levels(book: dict[str, Any]) -> list[dict[str, Any]]:
        by_price: dict[float, int] = {}
        for level in book.get("bids", []):
            try:
                if isinstance(level, dict):
                    p = float(level["price"])
                    q = int(float(level.get("size", level.get("quantity", 0))))
                elif isinstance(level, (list, tuple)):
                    p, q = float(level[0]), int(float(level[1]))
                else:
                    continue
            except (TypeError, ValueError, KeyError):
                continue
            if q > 0 and 0.0 <= p <= 1.0:
                rp = round(p, 6)
                by_price[rp] = by_price.get(rp, 0) + q
        # best bids first
        return [{"price": p, "size": by_price[p]} for p in sorted(by_price, reverse=True)]

    @staticmethod
    def _best_ask(book: dict[str, Any]) -> float:
        prices = []
        for level in book.get("asks", []):
            try:
                p = float(level["price"]) if isinstance(level, dict) else float(level[0])
                if 0.0 <= p <= 1.0:
                    prices.append(p)
            except (TypeError, ValueError, KeyError, IndexError):
                pass
        return min(prices) if prices else 1.0

    @staticmethod
    def _best_bid(book: dict[str, Any]) -> float:
        prices = []
        for level in book.get("bids", []):
            try:
                p = float(level["price"]) if isinstance(level, dict) else float(level[0])
                if 0.0 <= p <= 1.0:
                    prices.append(p)
            except (TypeError, ValueError, KeyError, IndexError):
                pass
        return max(prices) if prices else 0.0

    def get_fee_rate(self, token_id: str) -> float:
        """Return the taker fee rate for a Polymarket token.

        Checks feesEnabled via the Gamma API. If fees are disabled (e.g. NHL),
        returns 0.0. If enabled, returns the fixed sports taker rate of 0.03.

        The CLOB fee-rate endpoint returns incorrect values (1000 bps instead of
        300 bps for sports), so we do not use it for the rate itself.
        """
        if not token_id:
            raise RuntimeError("Polymarket fee-rate lookup failed: empty token_id")
        res = self._session.get(
            f"{self.gamma_host}/markets",
            params={"clob_token_ids": token_id},
            timeout=_TIMEOUT,
        )
        res.raise_for_status()
        data = res.json()
        if not data:
            raise RuntimeError(f"Polymarket gamma API returned no market for token_id={token_id}")
        fees_enabled = data[0].get("feesEnabled", True)
        return 0.03 if fees_enabled else 0.0

    def get_question_by_token(self, token_id: str) -> str:
        """Fetch the market question text for a Polymarket token ID via Gamma API.

        Returns empty string on failure rather than raising.
        """
        if not token_id:
            return ""
        try:
            res = self._session.get(
                f"{self.gamma_host}/markets",
                params={"clob_token_ids": token_id},
                timeout=_TIMEOUT,
            )
            res.raise_for_status()
            payload = res.json()
            market = payload[0] if isinstance(payload, list) and payload else payload
            if isinstance(market, dict):
                return str(market.get("question") or "").strip()
        except Exception:
            pass
        return ""

    def _fetch_book(self, token_id: str) -> dict[str, Any]:
        res = self._session.get(
            f"{self.clob_host}/book",
            params={"token_id": token_id},
            timeout=_TIMEOUT,
        )
        res.raise_for_status()
        return res.json()

    def get_quotes(
        self,
        yes_token_id: str,
        no_token_id: str,
        market_slug: str = "",
    ) -> dict[str, Any]:
        """Fetch and normalise orderbook quotes for both YES and NO tokens.

        If token IDs are empty, resolves them from market_slug via Gamma API.

        Returns:
            yes_bid, yes_ask, no_bid, no_ask (floats in [0,1])
            depth.yes_asks  — ask levels for buying YES  [{price, size}, ...]
            depth.no_asks   — ask levels for buying NO   [{price, size}, ...]
            yes_token_id, no_token_id (resolved)
        """
        yid = yes_token_id.strip()
        nid = no_token_id.strip()
        if not yid or not nid:
            if not market_slug:
                raise RuntimeError(
                    "Polymarket: need yes_token_id + no_token_id, or market_slug for auto-resolution"
                )
            yid, nid = self._resolve_tokens(market_slug)

        # Fetch YES and NO orderbooks in parallel — halves per-pair Polymarket latency
        with ThreadPoolExecutor(max_workers=2) as pool:
            yes_fut = pool.submit(self._fetch_book, yid)
            no_fut = pool.submit(self._fetch_book, nid)
            yes_book = yes_fut.result()
            no_book = no_fut.result()

        return {
            "yes_bid": self._best_bid(yes_book),
            "yes_ask": self._best_ask(yes_book),
            "no_bid": self._best_bid(no_book),
            "no_ask": self._best_ask(no_book),
            "depth": {
                "yes_asks": self._parse_ask_levels(yes_book),
                "no_asks": self._parse_ask_levels(no_book),
                "yes_bids": self._parse_bid_levels(yes_book),
                "no_bids": self._parse_bid_levels(no_book),
            },
            "yes_token_id": yid,
            "no_token_id": nid,
        }

    def place_order(
        self, token_id: str, side: str, size: int
    ) -> dict[str, Any]:
        """Place a FAK sweep order on Polymarket International (V2).

        FAK (Fill and Kill) fills whatever is immediately available at ≤0.99
        and cancels the remainder — never posts to the book. Accepts partial
        fills. size must be an integer contract count.
        Raises if 0 contracts were filled.
        """
        self._ensure_client()
        from py_clob_client_v2 import OrderArgs, OrderType, Side  # type: ignore

        side_uc = side.strip().upper()
        if side_uc not in {"BUY", "SELL"}:
            raise ValueError("Polymarket side must be 'buy' or 'sell'")

        side_enum = Side.BUY if side_uc == "BUY" else Side.SELL

        resp = self._client.create_and_post_order(
            order_args=OrderArgs(
                token_id=token_id,
                price=0.99,
                size=int(size),
                side=side_enum,
            ),
            order_type=OrderType.FAK,
        )

        taking = resp.get("takingAmount", "0") or "0"
        if not taking or taking == "0":
            raise RuntimeError(
                f"Polymarket GTC order filled 0 contracts "
                f"(status={resp.get('status','?')}, errorMsg={resp.get('errorMsg','')})"
            )
        return resp


# ── Excel / CSV pairs loader ──────────────────────────────────────────────────

# Polymarket fee rates must be fetched dynamically via the fee-rate endpoint.
# Do not hardcode category rates here; compliance requires using:
#   GET https://clob.polymarket.com/fee-rate?token_id={token_id}


def _pick(row: dict, *keys: str, default: str = "") -> str:
    """Return the first non-empty string value found among the given keys."""
    for key in keys:
        v = row.get(key)
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    return default


def _parse_json_list(value: Any) -> list[str]:
    """Parse a JSON list or comma-delimited string into a list of strings."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    raw = str(value).strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(v).strip() for v in parsed if str(v).strip()]
    except json.JSONDecodeError:
        pass
    return [v.strip() for v in raw.split(",") if v.strip()]


def get_latest_pairs_file(input_dir: str | None) -> str | None:
    """Return the path to the newest .csv or .xlsx/.xls file in the input_files directory."""
    if not input_dir:
        return None
    p = Path(input_dir)
    if not p.exists() or not p.is_dir():
        return None
    candidates = [
        f for ext in ("*.csv", "*.xlsx", "*.xls") for f in p.glob(ext)
    ]
    if not candidates:
        return None
    return str(max(candidates, key=lambda f: f.stat().st_mtime))


def load_pairs(path: str) -> list[dict[str, Any]]:
    """Load the manual pairs list from an Excel (.xlsx) or CSV file.

    Accepts the actual column names used in Pairs_for_Kalshi_and_Polymarket.xlsx
    as well as the generic fallback names:

        Kalshi ticker  : kalshi_market_id  | kalshi_ticker
        Polymarket slug: poly_slug         | polymarket_market_slug
        YES token ID   : polymarket_yes_token_id   (optional — resolved via slug at scan time)
        NO  token ID   : polymarket_no_token_id    (optional)
        Title          : title_clean       | title
        Category       : category_tag      | category
        Resolution date: time to expiration (2 months out) | resolution_time_utc | resolution_date
        Optional URLs  : kalshi_url | poly_url | poly_event_url
        Active flag    : active  (optional — if absent all rows are treated as active)
        Row ID         : pair_id

    Returns a list of dicts for active pairs only.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Pairs file not found: {path}")

    rows: list[dict[str, Any]] = []

    if p.suffix.lower() in {".xlsx", ".xls"}:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [
            str(cell.value or "").strip().lower()
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                raw_headers[i]: row[i]
                for i in range(min(len(raw_headers), len(row)))
            })
        wb.close()
    else:
        import csv
        with p.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = [{k.strip().lower(): v for k, v in r.items()} for r in reader]

    has_active_col = any("active" in (row or {}) for row in rows[:5])

    pairs: list[dict[str, Any]] = []
    for row in rows:
        # Respect an explicit active column if present; otherwise include all rows
        if has_active_col:
            active = str(row.get("active", "true") or "true").strip().lower()
            if active in {"false", "0", "no"}:
                continue

        pair_id = _pick(row, "pair_id", "poly_market_id")
        # Accept both "kalshi_market_id" (actual file) and "kalshi_ticker" (generic)
        kalshi_ticker = _pick(row, "kalshi_market_id", "kalshi_ticker")
        if not pair_id or not kalshi_ticker:
            continue

        pairs.append({
            "pair_id": pair_id,
            "title": _pick(row, "title_clean", "title", "poly_title", "kalshi_title", default=pair_id),
            "kalshi_ticker": kalshi_ticker,
            "kalshi_ticker_b": _pick(row, "kalshi_market_id_b", "kalshi_ticker_b", default=""),
            # Token IDs are optional — if absent they are resolved via slug at scan time
            "polymarket_yes_token_id": _pick(row, "polymarket_yes_token_id"),
            "polymarket_no_token_id": _pick(row, "polymarket_no_token_id"),
            # Accept both "poly_slug" (actual file) and "polymarket_market_slug" (generic)
            "polymarket_market_slug": _pick(row, "poly_slug", "polymarket_market_slug"),
            "poly_outcomes": _parse_json_list(_pick(row, "poly_outcomes_json", "poly_outcomes")),
            "poly_token_ids": _parse_json_list(_pick(row, "poly_token_ids_json", "poly_token_ids")),
            "poly_primary_outcome": _pick(row, "poly_primary_outcome", "poly_outcome"),
            "poly_event_url": _pick(row, "poly_event_url"),
            "polymarket_url": _pick(row, "poly_url", "polymarket_url"),
            "kalshi_url": _pick(row, "kalshi_url"),
            # Accept several resolution-date column name variants
            "resolution_date": _pick(
                row,
                "time to expiration (2 months out)",
                "resolution_time_utc",
                "resolution_date",
                "resolution_time",
            ),
            # Accept both "category_tag" (actual file) and "category" (generic)
            "category": _pick(row, "category_tag", "category", default="default"),
        })
        pairs[-1]["poly_fee_rate"] = None

    # ── Data validation ──
    seen_ids: set[str] = set()
    validated: list[dict[str, Any]] = []
    for p in pairs:
        pid = p["pair_id"]
        # Duplicate pair_id check
        if pid in seen_ids:
            print(f"  [WARN] Duplicate pair_id '{pid}' — skipping duplicate row")
            continue
        seen_ids.add(pid)

        # Kalshi ticker format: should be non-empty, alphanumeric + hyphens
        kt = p["kalshi_ticker"]
        if not all(ch.isalnum() or ch in "-_" for ch in kt):
            print(f"  [WARN] pair {pid}: unusual kalshi_ticker format '{kt}'")

        # Polymarket slug check
        if not p.get("polymarket_market_slug"):
            print(f"  [WARN] pair {pid}: empty polymarket_market_slug — tokens will need resolution")

        # Multi-outcome mapping sanity (requires explicit primary outcome)
        outcomes = [str(o).strip().lower() for o in p.get("poly_outcomes", [])]
        if len(outcomes) > 2:
            primary = str(p.get("poly_primary_outcome", "")).strip().lower()
            if not primary:
                print(f"  [WARN] pair {pid}: missing poly_primary_outcome for multi-outcome market")
            elif primary not in outcomes:
                print(f"  [WARN] pair {pid}: poly_primary_outcome '{primary}' not in outcomes {outcomes}")

        # Resolution date format check
        rd = p.get("resolution_date")
        if rd and isinstance(rd, str):
            rd_str = rd.strip()[:10]
            try:
                from datetime import date as _date
                _date.fromisoformat(rd_str)
            except (ValueError, TypeError):
                print(f"  [WARN] pair {pid}: invalid resolution_date format '{rd}' — will default to 365 days")
                p["resolution_date"] = ""

        validated.append(p)

    return validated
