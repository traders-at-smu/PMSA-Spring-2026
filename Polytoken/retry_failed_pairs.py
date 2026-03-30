"""retry_failed_pairs.py — Re-validate pairs in failed_pairs.json and remove those that pass.

Usage:
    python retry_failed_pairs.py                         # dry-run: show what would be cleared
    python retry_failed_pairs.py --apply                 # re-validate and rewrite failed_pairs.json
    python retry_failed_pairs.py --clear-all             # remove ALL entries without re-validating
    python retry_failed_pairs.py --clear-id PAIR_001     # remove a specific pair_id
    python retry_failed_pairs.py --pairs /path/pairs.xlsx --failed /path/failed.json

The script reads the failed_pairs.json log (one JSON object per line), looks up each
pair_id in the Excel pairs file to get the Kalshi ticker and Polymarket slug, then
re-runs the same API checks that the bot uses. If a pair now passes it is removed from
the failed log so the bot will attempt it again on the next run.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import requests

# ── Default path resolution ───────────────────────────────────────────────────

V2_CONTRACTS_FILENAME = "Pairs_for_Kalshi_and_Polymarket.xlsx"
KALSHI_BASE = "https://api.elections.kalshi.com/trade-api/v2"
POLY_GAMMA_BASE = "https://gamma-api.polymarket.com"
POLY_CLOB_BASE = "https://clob.polymarket.com"
HTTP_TIMEOUT = 15


def _resolve_default_pairs_path() -> Path:
    script_path = Path(__file__).resolve()
    for base in script_path.parents:
        candidate = base / "V5" / V2_CONTRACTS_FILENAME
        if candidate.exists():
            return candidate
    return script_path.parents[1] / "V5" / V2_CONTRACTS_FILENAME


def _resolve_default_failed_path() -> Path:
    script_path = Path(__file__).resolve()
    for base in script_path.parents:
        candidate = base / "V5" / "failed_pairs.json"
        if candidate.exists():
            return candidate
    return script_path.parents[1] / "V5" / "failed_pairs.json"


DEFAULT_PAIRS_PATH = _resolve_default_pairs_path()
DEFAULT_FAILED_PATH = _resolve_default_failed_path()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _load_failed_entries(path: Path) -> list[dict[str, Any]]:
    """Load all entries from failed_pairs.json (JSONL format)."""
    entries: list[dict[str, Any]] = []
    if not path.exists():
        return entries
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                pass
    return entries


def _load_pairs_lookup(pairs_path: Path) -> dict[str, dict[str, str]]:
    """Return {pair_id: {kalshi_ticker, poly_slug}} from the Excel or CSV pairs file."""
    lookup: dict[str, dict[str, str]] = {}
    if not pairs_path.exists():
        return lookup

    if pairs_path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("WARNING: openpyxl not installed — cannot read .xlsx pairs file", file=sys.stderr)
            return lookup
        wb = load_workbook(pairs_path, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [
            str(cell.value or "").strip().lower()
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {raw_headers[i]: row[i] for i in range(min(len(raw_headers), len(row)))}
            pair_id = str(row_dict.get("pair_id") or "").strip()
            kalshi = str(row_dict.get("kalshi_market_id") or row_dict.get("kalshi_ticker") or "").strip()
            slug = str(row_dict.get("poly_slug") or row_dict.get("polymarket_market_slug") or "").strip()
            if pair_id:
                lookup[pair_id] = {"kalshi_ticker": kalshi, "poly_slug": slug}
        wb.close()
    else:
        import csv
        with pairs_path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                row = {k.strip().lower(): v for k, v in row.items()}
                pair_id = str(row.get("pair_id") or "").strip()
                kalshi = str(row.get("kalshi_market_id") or row.get("kalshi_ticker") or "").strip()
                slug = str(row.get("poly_slug") or row.get("polymarket_market_slug") or "").strip()
                if pair_id:
                    lookup[pair_id] = {"kalshi_ticker": kalshi, "poly_slug": slug}
    return lookup


def _check_kalshi(ticker: str) -> str | None:
    """Return None if OK, or an error message string if the market is not tradeable."""
    try:
        res = requests.get(f"{KALSHI_BASE}/markets/{ticker}", timeout=HTTP_TIMEOUT)
        res.raise_for_status()
        market = res.json().get("market", {})
        status = str(market.get("status", "")).strip().lower()
        if status not in {"open", "active"}:
            return f"Kalshi market status is '{status}'"
        return None
    except Exception as exc:
        return str(exc)


def _check_polymarket(slug: str) -> str | None:
    """Return None if OK, or an error message string if the market is not tradeable."""
    try:
        res = requests.get(f"{POLY_GAMMA_BASE}/markets", params={"slug": slug}, timeout=HTTP_TIMEOUT)
        res.raise_for_status()
        payload = res.json()
        if not isinstance(payload, list) or not payload:
            return f"No market found for slug '{slug}'"
        market = payload[0]

        import json as _json
        raw_ids = market.get("clobTokenIds", [])
        if isinstance(raw_ids, str):
            raw_ids = _json.loads(raw_ids)
        if len(raw_ids) < 2:
            return f"fewer than 2 CLOB token IDs"

        outcomes = [o.strip().lower() for o in (market.get("outcomes") or [])]
        if "yes" not in outcomes or "no" not in outcomes:
            return f"non-binary outcomes {outcomes!r}"

        yes_id = str(raw_ids[outcomes.index("yes")]).strip()
        no_id = str(raw_ids[outcomes.index("no")]).strip()
        if yes_id == no_id:
            return f"YES and NO token IDs are identical"

        # Spot-check CLOB orderbook accessibility for YES token
        book_res = requests.get(f"{POLY_CLOB_BASE}/book", params={"token_id": yes_id}, timeout=HTTP_TIMEOUT)
        book_res.raise_for_status()
        return None
    except Exception as exc:
        return str(exc)


# ── Core logic ────────────────────────────────────────────────────────────────

def retry_failed_pairs(
    failed_path: Path,
    pairs_path: Path,
    apply: bool,
    clear_all: bool,
    clear_ids: set[str],
) -> None:
    entries = _load_failed_entries(failed_path)
    if not entries:
        print(f"No entries in {failed_path}")
        return

    # Deduplicate by pair_id — keep only the most recent entry per pair
    seen: dict[str, dict[str, Any]] = {}
    for e in entries:
        pid = e.get("pair_id", "")
        seen[pid] = e  # last wins = most recent (entries are appended chronologically)
    unique_entries = list(seen.values())

    print(f"Failed pairs log: {failed_path}")
    print(f"Pairs file:       {pairs_path}")
    print(f"Unique failed pairs: {len(unique_entries)}")
    if clear_all:
        print("Mode: --clear-all (skip re-validation, remove everything)")
    elif clear_ids:
        print(f"Mode: --clear-id {clear_ids}")
    else:
        print("Mode: re-validate each pair against live APIs")
    print()

    if clear_all:
        if apply:
            failed_path.write_text("", encoding="utf-8")
            print(f"Cleared all {len(unique_entries)} entries from {failed_path}")
        else:
            print(f"Dry-run: would remove all {len(unique_entries)} entries.")
            print("Pass --apply to write the change.")
        return

    if clear_ids:
        keep = [e for e in unique_entries if e.get("pair_id") not in clear_ids]
        removed = len(unique_entries) - len(keep)
        if apply:
            with failed_path.open("w", encoding="utf-8") as f:
                for e in keep:
                    f.write(json.dumps(e) + "\n")
            print(f"Removed {removed} entry/entries. {len(keep)} remain.")
        else:
            for e in unique_entries:
                if e.get("pair_id") in clear_ids:
                    print(f"  would remove: {e['pair_id']}")
            print(f"\nDry-run: would remove {removed} entry/entries. Pass --apply.")
        return

    # Re-validate mode
    pairs_lookup = _load_pairs_lookup(pairs_path)
    if not pairs_lookup:
        print("WARNING: could not load pairs file — re-validation will be skipped for unknown pairs")

    keep: list[dict[str, Any]] = []
    cleared: list[str] = []
    still_failing: list[tuple[str, str]] = []

    for entry in unique_entries:
        pid = entry.get("pair_id", "?")
        pair_info = pairs_lookup.get(pid)

        if pair_info is None:
            print(f"  [SKIP]  {pid:<35} — not found in pairs file (pair may have been removed)")
            keep.append(entry)
            continue

        kalshi_ticker = pair_info.get("kalshi_ticker", "")
        poly_slug = pair_info.get("poly_slug", "")

        if not kalshi_ticker or not poly_slug:
            print(f"  [SKIP]  {pid:<35} — missing ticker or slug in pairs file")
            keep.append(entry)
            continue

        # Re-run the checks
        kalshi_err = _check_kalshi(kalshi_ticker)
        poly_err = _check_polymarket(poly_slug)

        if kalshi_err is None and poly_err is None:
            print(f"  [PASS]  {pid:<35} — {kalshi_ticker} / {poly_slug}")
            cleared.append(pid)
        else:
            reasons = []
            if kalshi_err:
                reasons.append(f"Kalshi: {kalshi_err}")
            if poly_err:
                reasons.append(f"Polymarket: {poly_err}")
            reason_str = "; ".join(reasons)
            print(f"  [FAIL]  {pid:<35} — {reason_str}")
            still_failing.append((pid, reason_str))
            keep.append(entry)

    print()
    print(f"Summary: {len(cleared)} now pass, {len(still_failing)} still failing, "
          f"{len(unique_entries) - len(cleared) - len(still_failing)} skipped")

    if not cleared:
        print("Nothing to clear.")
        return

    if not apply:
        print(f"\nDry-run: would remove {len(cleared)} entry/entries from failed log. Pass --apply.")
        return

    with failed_path.open("w", encoding="utf-8") as f:
        for e in keep:
            f.write(json.dumps(e) + "\n")
    print(f"\nUpdated {failed_path} — removed {len(cleared)} entry/entries, {len(keep)} remain.")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Re-validate failed pairs and remove those that now pass from failed_pairs.json."
    )
    parser.add_argument(
        "--failed",
        default=str(DEFAULT_FAILED_PATH),
        help=f"Path to failed_pairs.json (default: {DEFAULT_FAILED_PATH})",
    )
    parser.add_argument(
        "--pairs",
        default=str(DEFAULT_PAIRS_PATH),
        help=f"Path to pairs .xlsx file (default: {DEFAULT_PAIRS_PATH})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write changes to disk. Without this flag the script runs in dry-run mode.",
    )
    parser.add_argument(
        "--clear-all",
        action="store_true",
        help="Remove ALL entries from failed_pairs.json without re-validating.",
    )
    parser.add_argument(
        "--clear-id",
        metavar="PAIR_ID",
        action="append",
        default=[],
        help="Remove a specific pair_id. Can be specified multiple times.",
    )
    args = parser.parse_args()

    retry_failed_pairs(
        failed_path=Path(args.failed).resolve(),
        pairs_path=Path(args.pairs).resolve(),
        apply=args.apply,
        clear_all=args.clear_all,
        clear_ids=set(args.clear_id),
    )


if __name__ == "__main__":
    main()
