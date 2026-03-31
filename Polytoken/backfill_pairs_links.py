"""backfill_pairs_links.py — Backfill URLs and outcome mapping columns in pairs Excel.

Usage:
    python backfill_pairs_links.py                      # dry-run (no write)
    python backfill_pairs_links.py --apply              # write updates
    python backfill_pairs_links.py --target /path/file  # explicit pairs file
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import requests


V2_CONTRACTS_FILENAME = "Pairs_for_Kalshi_and_Polymarket.xlsx"
KALSHI_BASE = "https://api.elections.kalshi.com/trade-api/v2"
POLY_GAMMA_BASE = "https://gamma-api.polymarket.com"
HTTP_TIMEOUT = 15


def _resolve_default_path() -> Path:
    script_path = Path(__file__).resolve()
    for base in script_path.parents:
        candidate = base / "V5" / V2_CONTRACTS_FILENAME
        if candidate.exists():
            return candidate
    return script_path.parents[1] / "V5" / V2_CONTRACTS_FILENAME


DEFAULT_PATH = _resolve_default_path()


def _normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _normalize_outcome_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    cleaned = []
    for ch in text:
        cleaned.append(ch if ch.isalnum() else " ")
    return " ".join("".join(cleaned).split())


def _match_primary_outcome(outcomes: list[str], kalshi_label: str) -> str:
    norm_k = _normalize_outcome_label(kalshi_label)
    if not outcomes:
        return ""
    matches = [o for o in outcomes if _normalize_outcome_label(o) == norm_k]
    if len(matches) == 1:
        return matches[0]
    fuzzy = [
        o for o in outcomes
        if norm_k and (norm_k in _normalize_outcome_label(o) or _normalize_outcome_label(o) in norm_k)
    ]
    if len(fuzzy) == 1:
        return fuzzy[0]
    return ""


def _slugify(text: str) -> str:
    text = text.lower().strip()
    slug: list[str] = []
    prev_dash = False
    for ch in text:
        is_alnum = "a" <= ch <= "z" or "0" <= ch <= "9"
        if is_alnum:
            slug.append(ch)
            prev_dash = False
        elif not prev_dash:
            slug.append("-")
            prev_dash = True
    return "".join(slug).strip("-")


def _fetch_kalshi_market(ticker: str) -> dict[str, Any] | None:
    if not ticker:
        return None
    res = requests.get(f"{KALSHI_BASE}/markets/{ticker}", timeout=HTTP_TIMEOUT)
    if res.status_code == 404:
        return None
    res.raise_for_status()
    return res.json().get("market", {})


def _build_kalshi_url(kalshi_id: str) -> str:
    """Return the best navigable Kalshi URL for a market ticker.

    Tries to build the full 3-segment public URL:
      https://kalshi.com/markets/{event_ticker}/{event_slug}/{market_ticker}
    Falls back to the event-page URL (2-segment) on API failure, then to a
    bare market-ticker URL as a last resort.
    """
    if not kalshi_id:
        return ""
    market = _fetch_kalshi_market(kalshi_id)
    if not market:
        # Last resort: use the ticker itself (may not navigate, but beats empty)
        return f"https://kalshi.com/markets/{kalshi_id.lower()}"
    event_ticker = str(market.get("event_ticker") or "").strip().lower()
    if not event_ticker:
        event_ticker = kalshi_id.lower()
    base_event_url = f"https://kalshi.com/markets/{event_ticker}"
    # Try to fetch the event to get its slug for a full 3-segment URL
    try:
        r = requests.get(f"{KALSHI_BASE}/events/{event_ticker.upper()}", timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        event = r.json().get("event", {})
        title = str(event.get("title") or "").strip()
        if title:
            slug = _slugify(title)
            if slug:
                return f"https://kalshi.com/markets/{event_ticker}/{slug}/{kalshi_id.lower()}"
    except Exception:
        pass
    return base_event_url


def _fetch_poly_market(slug: str) -> dict[str, Any] | None:
    if not slug:
        return None
    res = requests.get(f"{POLY_GAMMA_BASE}/markets", params={"slug": slug}, timeout=HTTP_TIMEOUT)
    res.raise_for_status()
    payload = res.json()
    if isinstance(payload, list) and payload:
        return payload[0]
    if isinstance(payload, dict) and payload.get("conditionId"):
        return payload
    return None


def _extract_outcomes_tokens(market: dict[str, Any]) -> tuple[list[str], list[str]]:
    raw_ids = market.get("clobTokenIds", [])
    if isinstance(raw_ids, str):
        try:
            raw_ids = json.loads(raw_ids)
        except json.JSONDecodeError:
            raw_ids = [raw_ids]
    outcomes = market.get("outcomes") or []
    if isinstance(outcomes, str):
        try:
            outcomes = json.loads(outcomes)
        except json.JSONDecodeError:
            outcomes = [outcomes]
    outcomes = [str(o).strip() for o in outcomes if str(o).strip()]
    token_ids = [str(t).strip() for t in raw_ids if str(t).strip()]
    return outcomes, token_ids


def backfill(target_path: Path, apply: bool) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not target_path.exists():
        print(f"ERROR: pairs file not found: {target_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(target_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map: dict[str, int] = {
        _normalize_header(h): i + 1
        for i, h in enumerate(headers)
        if _normalize_header(h)
    }

    def ensure_col(field: str) -> int:
        key = _normalize_header(field)
        col = header_map.get(key)
        if col:
            return col
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=field)
        header_map[key] = col
        return col

    # Ensure required columns exist
    col_pair_id = ensure_col("pair_id")
    col_poly_slug = ensure_col("poly_slug")
    col_kalshi_id = ensure_col("kalshi_market_id")
    col_poly_url = ensure_col("poly_url")
    col_kalshi_url = ensure_col("kalshi_url")
    col_poly_event_url = ensure_col("poly_event_url")
    col_outcomes = ensure_col("poly_outcomes_json")
    col_token_ids = ensure_col("poly_token_ids_json")
    col_primary = ensure_col("poly_primary_outcome")

    updates = 0
    for r in range(2, ws.max_row + 1):
        pair_id = str(ws.cell(row=r, column=col_pair_id).value or "").strip()
        poly_slug = str(ws.cell(row=r, column=col_poly_slug).value or "").strip()
        kalshi_id = str(ws.cell(row=r, column=col_kalshi_id).value or "").strip()
        if not poly_slug and not kalshi_id:
            continue

        new_poly_url = f"https://polymarket.com/market/{poly_slug}" if poly_slug else ""
        new_kalshi_url = _build_kalshi_url(kalshi_id)

        poly_market = _fetch_poly_market(poly_slug) if poly_slug else None
        outcomes, token_ids = _extract_outcomes_tokens(poly_market) if poly_market else ([], [])
        event_slug = str(poly_market.get("eventSlug") or poly_market.get("event_slug") or "").strip() if poly_market else ""
        new_event_url = f"https://polymarket.com/event/{event_slug}" if event_slug else ""

        primary_outcome = str(ws.cell(row=r, column=col_primary).value or "").strip()
        if outcomes:
            outcomes_lc = [o.strip().lower() for o in outcomes]
            if not primary_outcome:
                if len(outcomes_lc) == 2 and "yes" in outcomes_lc and "no" in outcomes_lc:
                    primary_outcome = "yes"
                else:
                    # _build_kalshi_url already fetched the market; fetch again only for the label
                    kalshi_market = _fetch_kalshi_market(kalshi_id) if kalshi_id else None
                    kalshi_label = ""
                    if kalshi_market:
                        kalshi_label = str(kalshi_market.get("yes_sub_title") or kalshi_market.get("title") or "").strip()
                    primary_outcome = _match_primary_outcome(outcomes, kalshi_label)

        def maybe_set(col: int, new_val: str) -> None:
            nonlocal updates
            old_val = str(ws.cell(row=r, column=col).value or "").strip()
            if new_val != old_val:
                updates += 1
                if apply:
                    ws.cell(row=r, column=col, value=new_val)
                print(f"  {'UPDATE' if apply else 'would update'} row {r} ({pair_id}) {old_val!r} -> {new_val!r}")

        maybe_set(col_poly_url, new_poly_url)
        maybe_set(col_kalshi_url, new_kalshi_url)
        if poly_market:
            maybe_set(col_outcomes, json.dumps(outcomes))
            maybe_set(col_token_ids, json.dumps(token_ids))
            if primary_outcome:
                maybe_set(col_primary, primary_outcome)
            if new_event_url:
                maybe_set(col_poly_event_url, new_event_url)

    if not apply:
        print(f"\nDry-run complete. {updates} change(s) would be applied.")
        wb.close()
        return

    wb.save(target_path)
    wb.close()
    print(f"\nSaved {target_path} — applied {updates} change(s).")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill URLs and outcome mapping columns in the pairs Excel file."
    )
    parser.add_argument(
        "--target",
        default=str(DEFAULT_PATH),
        help=f"Path to the pairs .xlsx file (default: {DEFAULT_PATH})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write updates to the file (default: dry-run).",
    )
    args = parser.parse_args()
    backfill(Path(args.target).resolve(), apply=args.apply)


if __name__ == "__main__":
    main()
