"""backfill_category.py — Backfill the category_tag column from Polymarket tag metadata.

Targets:
  - V5/Pairs_for_Kalshi_and_Polymarket.xlsx  (master Excel)
  - Polytoken/pairs/*/pairs.csv              (per-user staging CSVs)

Usage:
    python backfill_category.py                      # dry-run (no write)
    python backfill_category.py --apply              # write updates
    python backfill_category.py --target /path/file  # explicit Excel path only
    python backfill_category.py --csv-only           # only update staging CSVs
    python backfill_category.py --xlsx-only          # only update master Excel
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from pathlib import Path
from typing import Any

import requests


POLY_GAMMA_BASE = "https://gamma-api.polymarket.com"
HTTP_TIMEOUT = 15
REQUEST_DELAY = 0.2  # seconds between API calls to avoid rate limiting


def _resolve_default_xlsx() -> Path:
    script_path = Path(__file__).resolve()
    for base in script_path.parents:
        candidate = base / "V5" / "Pairs_for_Kalshi_and_Polymarket.xlsx"
        if candidate.exists():
            return candidate
    return script_path.parents[1] / "V5" / "Pairs_for_Kalshi_and_Polymarket.xlsx"


USER_PAIRS_DIR = Path(__file__).resolve().parent / "pairs"
DEFAULT_XLSX = _resolve_default_xlsx()


def _normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _fetch_poly_market(slug: str) -> dict[str, Any] | None:
    if not slug:
        return None
    try:
        res = requests.get(
            f"{POLY_GAMMA_BASE}/markets",
            params={"slug": slug},
            timeout=HTTP_TIMEOUT,
        )
        res.raise_for_status()
        payload = res.json()
        if isinstance(payload, list) and payload:
            return payload[0]
        if isinstance(payload, dict) and payload.get("conditionId"):
            return payload
    except Exception as exc:
        print(f"  Warning: failed to fetch market '{slug}': {exc}", file=sys.stderr)
    return None


def _fetch_event_tags(market: dict[str, Any]) -> list:
    """Fetch tags from the parent event. Tags live on the event, not the market."""
    events = market.get("events") or []
    event_slug = ""
    if isinstance(events, list) and events and isinstance(events[0], dict):
        event_slug = str(events[0].get("slug") or "").strip()
    if not event_slug:
        return []
    try:
        res = requests.get(
            f"{POLY_GAMMA_BASE}/events",
            params={"slug": event_slug},
            timeout=HTTP_TIMEOUT,
        )
        res.raise_for_status()
        event_data = res.json()
        if isinstance(event_data, list) and event_data:
            event_data = event_data[0]
        if isinstance(event_data, dict):
            return event_data.get("tags") or []
    except Exception:
        pass
    return []


def _extract_category(market: dict[str, Any]) -> str:
    tags = _fetch_event_tags(market)
    for tag in tags:
        if isinstance(tag, dict):
            label = str(tag.get("label") or tag.get("slug") or "").strip().lower()
            if label:
                return label
    return "default"


# ---------------------------------------------------------------------------
# Excel backfill
# ---------------------------------------------------------------------------

def backfill_xlsx(target_path: Path, apply: bool) -> int:
    """Backfill category_tag in the master Excel file. Returns number of changes."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not target_path.exists():
        print(f"Skipping Excel (not found): {target_path}", file=sys.stderr)
        return 0

    wb = load_workbook(target_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map: dict[str, int] = {
        _normalize_header(h): i + 1
        for i, h in enumerate(headers)
        if _normalize_header(h)
    }

    col_slug = header_map.get("polyslug") or header_map.get("poly_slug")
    col_category = header_map.get("categorytag") or header_map.get("category_tag")
    col_pair_id = header_map.get("pairid") or header_map.get("pair_id")

    if col_slug is None:
        print("ERROR: could not find poly_slug column in Excel", file=sys.stderr)
        wb.close()
        return 0

    # Add category_tag column if missing
    if col_category is None:
        col_category = ws.max_column + 1
        ws.cell(row=1, column=col_category, value="category_tag")
        print(f"  Added missing category_tag column at col {col_category}")

    updates = 0
    for r in range(2, ws.max_row + 1):
        poly_slug = str(ws.cell(row=r, column=col_slug).value or "").strip()
        if not poly_slug:
            continue

        pair_id = str(ws.cell(row=r, column=col_pair_id).value or r) if col_pair_id else str(r)
        current = str(ws.cell(row=r, column=col_category).value or "").strip()

        # Skip rows that already have a non-default, non-empty category
        if current and current != "default":
            continue

        market = _fetch_poly_market(poly_slug)
        time.sleep(REQUEST_DELAY)
        if not market:
            continue

        new_category = _extract_category(market)
        if new_category == current:
            continue

        updates += 1
        action = "UPDATE" if apply else "would update"
        print(f"  {action} row {r} ({pair_id}): {current!r} -> {new_category!r}")
        if apply:
            ws.cell(row=r, column=col_category, value=new_category)

    if not apply:
        print(f"\nDry-run complete (Excel). {updates} change(s) would be applied.")
        wb.close()
        return updates

    if updates:
        wb.save(target_path)
        print(f"\nSaved {target_path} — applied {updates} change(s).")
    else:
        print(f"\nNo changes needed in {target_path}.")
    wb.close()
    return updates


# ---------------------------------------------------------------------------
# CSV backfill
# ---------------------------------------------------------------------------

def backfill_csv(csv_path: Path, apply: bool) -> int:
    """Backfill category_tag in a single pairs CSV. Returns number of changes."""
    if not csv_path.exists():
        return 0

    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return 0
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    if not rows:
        return 0

    if "category_tag" not in fieldnames:
        for anchor in ("title_clean", "pair_id"):
            if anchor in fieldnames:
                fieldnames.insert(fieldnames.index(anchor) + 1, "category_tag")
                break
        else:
            fieldnames.append("category_tag")
        print(f"  Added missing category_tag column to {csv_path.name}")

    updates = 0
    for row in rows:
        poly_slug = str(row.get("poly_slug") or "").strip()
        if not poly_slug:
            continue

        current = str(row.get("category_tag") or "").strip()
        if current and current != "default":
            continue

        pair_id = row.get("pair_id", "?")
        market = _fetch_poly_market(poly_slug)
        time.sleep(REQUEST_DELAY)
        if not market:
            continue

        new_category = _extract_category(market)
        if new_category == current:
            continue

        updates += 1
        action = "UPDATE" if apply else "would update"
        print(f"  {action} {csv_path.name} ({pair_id}): {current!r} -> {new_category!r}")
        if apply:
            row["category_tag"] = new_category

    if not apply:
        print(f"  Dry-run ({csv_path.name}): {updates} change(s) would be applied.")
        return updates

    if updates:
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        print(f"  Saved {csv_path} — applied {updates} change(s).")
    return updates


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill category_tag from Polymarket tag metadata."
    )
    parser.add_argument(
        "--target",
        default=str(DEFAULT_XLSX),
        help=f"Path to the master .xlsx file (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write updates (default: dry-run).",
    )
    parser.add_argument(
        "--csv-only",
        action="store_true",
        help="Only update per-user staging CSVs, skip Excel.",
    )
    parser.add_argument(
        "--xlsx-only",
        action="store_true",
        help="Only update master Excel, skip CSVs.",
    )
    args = parser.parse_args()

    total = 0

    if not args.csv_only:
        print(f"\n=== Excel: {args.target} ===")
        total += backfill_xlsx(Path(args.target).resolve(), apply=args.apply)

    if not args.xlsx_only and USER_PAIRS_DIR.exists():
        csv_files = sorted(USER_PAIRS_DIR.glob("*/pairs.csv"))
        if csv_files:
            print(f"\n=== Staging CSVs ({len(csv_files)} file(s)) ===")
            for csv_path in csv_files:
                total += backfill_csv(csv_path, apply=args.apply)
        else:
            print("\nNo staging CSVs found.")

    action = "applied" if args.apply else "would apply"
    print(f"\nDone. Total changes {action}: {total}")


if __name__ == "__main__":
    main()
