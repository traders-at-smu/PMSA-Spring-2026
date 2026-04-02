import argparse
import csv
import datetime
import getpass
import json
import sys
from pathlib import Path
from urllib.parse import urlparse

import requests

# API Constants
KALSHI_BASE = "https://api.elections.kalshi.com/trade-api/v2"
POLY_GAMMA_BASE = "https://gamma-api.polymarket.com"
HTTP_TIMEOUT_SEC = 10
OUTPUT_FIELDS = [
    "pair_id",
    "title_clean",
    "category_tag",
    "similarity_score",
    "poly_market_id",
    "poly_slug",
    "poly_url",
    "poly_event_url",
    "poly_outcomes_json",
    "poly_token_ids_json",
    "poly_primary_outcome",
    "kalshi_market_id",
    "kalshi_url",
    "expiry_poly_utc",
    "expiry_kalshi_utc",
    "resolution_time_utc",
]

POLY_CLOB_HOST = "https://clob.polymarket.com"

COUNTER_PATH = Path(__file__).resolve().parent / "counter.txt"
REPO_ROOT = Path(__file__).resolve().parent.parent
# Per-user staging directory: Polytoken/pairs/{username}/pairs.csv
USER_PAIRS_DIR = Path(__file__).resolve().parent / "pairs"
# Only the active V5 master list is checked for duplicates
XLSX_PATHS = [
    REPO_ROOT / "V5" / "Pairs_for_Kalshi_and_Polymarket.xlsx",
]


def _normalize_header(value):
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def load_existing_pair_keys() -> set[tuple[str, str]]:
    """Load all (poly_market_id, kalshi_market_id) tuples from the master Excel and all user CSVs."""
    keys: set[tuple[str, str]] = set()
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Warning: openpyxl not installed, skipping duplicate check", file=sys.stderr)
        return keys

    for xlsx_path in XLSX_PATHS:
        if not xlsx_path.exists():
            continue
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            header_map = {_normalize_header(h): i for i, h in enumerate(headers)}
            poly_idx = header_map.get("polymarketid") or header_map.get("poly_market_id")
            kalshi_idx = header_map.get("kalshimarketid") or header_map.get("kalshi_market_id")
            if poly_idx is None or kalshi_idx is None:
                # Try normalized versions
                for k, v in header_map.items():
                    if "poly" in k and "id" in k:
                        poly_idx = v
                    if "kalshi" in k and ("id" in k or "ticker" in k):
                        kalshi_idx = v
            if poly_idx is None or kalshi_idx is None:
                wb.close()
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                poly_id = str(row[poly_idx] or "").strip() if poly_idx < len(row) else ""
                kalshi_id = str(row[kalshi_idx] or "").strip() if kalshi_idx < len(row) else ""
                if poly_id and kalshi_id:
                    keys.add((poly_id, kalshi_id))
            wb.close()
        except Exception as exc:
            print(f"Warning: could not read {xlsx_path.name}: {exc}", file=sys.stderr)

    # Also scan all per-user staging CSVs so duplicates across users are caught
    if USER_PAIRS_DIR.exists():
        for csv_file in sorted(USER_PAIRS_DIR.glob("*/pairs.csv")):
            try:
                with csv_file.open(newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        poly_id = str(row.get("poly_market_id") or "").strip()
                        kalshi_id = str(row.get("kalshi_market_id") or "").strip()
                        if poly_id and kalshi_id:
                            keys.add((poly_id, kalshi_id))
            except Exception as exc:
                print(f"Warning: could not read {csv_file}: {exc}", file=sys.stderr)

    return keys


def get_next_pair_id():
    if not COUNTER_PATH.exists():
        COUNTER_PATH.write_text("1", encoding="utf-8")
        return "pair-0001"
    count = int(COUNTER_PATH.read_text(encoding="utf-8").strip())
    count += 1
    COUNTER_PATH.write_text(str(count), encoding="utf-8")
    return f"pair-{count:04d}"

def extract_poly_slug(url):
    """Safely extracts slug from Polymarket URL."""
    path = urlparse(url).path
    parts = [p for p in path.split("/") if p]
    try:
        if "event" in parts:
            return parts[-1]
        return parts[-1]
    except Exception:
        return None


def _poly_display_name(market):
    """Human-readable Polymarket market label."""
    if not isinstance(market, dict):
        return "Unknown market"
    question = str(market.get("question", "")).strip()
    slug = str(market.get("slug", "")).strip()
    if question and slug:
        return f"{question} [{slug}]"
    if question:
        return question
    if slug:
        return slug
    return str(market.get("conditionId", "Unknown market"))


def resolve_polymarket_contracts(poly_slug):
    """
    Resolves Polymarket contracts from either:
    - direct market slug (/market/<slug>)
    - event slug (/event/<slug> or /sports/.../<slug>)
    """
    slug = str(poly_slug or "").strip()
    if not slug:
        raise RuntimeError("Polymarket slug is empty")

    market_resp = requests.get(
        f"{POLY_GAMMA_BASE}/markets",
        params={"slug": slug},
        timeout=HTTP_TIMEOUT_SEC,
    )
    market_resp.raise_for_status()
    market_rows = market_resp.json()
    if isinstance(market_rows, list) and market_rows:
        return [market_rows[0]]
    if isinstance(market_rows, dict) and market_rows.get("conditionId"):
        return [market_rows]

    event_resp = requests.get(
        f"{POLY_GAMMA_BASE}/events",
        params={"slug": slug},
        timeout=HTTP_TIMEOUT_SEC,
    )
    event_resp.raise_for_status()
    event_rows = event_resp.json()
    if not isinstance(event_rows, list) or not event_rows:
        raise RuntimeError(f"No Polymarket market or event found for slug: {slug}")

    event = event_rows[0]
    raw_markets = event.get("markets", [])
    markets = [m for m in raw_markets if isinstance(m, dict)]
    if not markets:
        raise RuntimeError(f"Event has no markets for slug: {slug}")

    active_markets = [m for m in markets if bool(m.get("active")) and not bool(m.get("closed"))]
    candidates = active_markets or markets

    return candidates


def resolve_polymarket_market(poly_slug, context_label=""):
    """Backward-compatible single-market resolver with interactive disambiguation."""
    candidates = resolve_polymarket_contracts(poly_slug)
    if len(candidates) == 1:
        return candidates[0]

    prefix = f"{context_label} " if context_label else ""
    print(f"\n{prefix}Multiple Polymarket markets found for this event. Pick one:")
    for i, market in enumerate(candidates, 1):
        print(f"{i}) {_poly_display_name(market)}")

    choice = int(input(f"\n{prefix}Enter Polymarket choice number: ")) - 1
    if choice < 0 or choice >= len(candidates):
        raise RuntimeError("Polymarket choice out of range")
    return candidates[choice]


def extract_kalshi_market_like(url_or_ticker):
    """Extracts Kalshi market path from URL (supports nested /markets/<event>/<market>)."""
    value = str(url_or_ticker or "").strip()
    if not value:
        return ""

    if "://" not in value and "kalshi.com" in value:
        value = f"https://{value}"

    parsed = urlparse(value)
    if parsed.scheme and parsed.netloc:
        parts = [p for p in parsed.path.split("/") if p]
        for i, part in enumerate(parts):
            if part.lower() == "markets" and i + 1 < len(parts):
                return "/".join(parts[i + 1 :]).strip()
        return ""

    # Fallback for direct ticker/event-like input.
    return value


def _slugify(text: str) -> str:
    text = text.lower().strip()
    slug: list[str] = []
    prev_dash = False
    for ch in text:
        is_alnum = "a" <= ch <= "z" or "0" <= ch <= "9"
        if is_alnum:
            slug.append(ch)
            prev_dash = False
        elif not prev_dash:
            slug.append("-")
            prev_dash = True
    return "".join(slug).strip("-")


def _build_kalshi_url(event_ticker: str, market_ticker: str) -> str:
    """Return the full 3-segment Kalshi public URL for a market.

    Format: https://kalshi.com/markets/{event_ticker}/{event_slug}/{market_ticker}
    Falls back to the 2-segment event page if the events API can't supply a slug.
    """
    et = (event_ticker or market_ticker).strip().lower()
    base = f"https://kalshi.com/markets/{et}"
    try:
        r = requests.get(f"{KALSHI_BASE}/events/{et.upper()}", timeout=HTTP_TIMEOUT_SEC)
        r.raise_for_status()
        event = r.json().get("event", {})
        title = str(event.get("title") or "").strip()
        if title:
            slug = _slugify(title)
            if slug and market_ticker:
                return f"https://kalshi.com/markets/{et}/{slug}/{market_ticker.lower()}"
            if slug:
                return f"https://kalshi.com/markets/{et}/{slug}"
    except Exception:
        pass
    return base


def event_ticker_from_market_like(value):
    """Best-effort conversion of market-like value to Kalshi event ticker.

    URL path format is {event_ticker}/{slug}/{market_ticker}, so the first
    segment is the event ticker.  For bare tickers (no "/") the API-returned
    event_ticker is preferred; this heuristic is the last resort.
    """
    ticker = str(value or "").strip()
    if not ticker:
        return ""
    if "/" in ticker:
        # First segment of the path is always the event ticker
        return ticker.split("/")[0].strip().upper()
    # Bare ticker — no reliable heuristic; return as-is (caller uses API first)
    return ticker.upper()


def _get_market_payload(market_like):
    """
    Resolves a market payload from possible market-like values.
    Handles nested URL path forms such as "<event>/<market>".
    """
    candidates = []
    for candidate in [market_like, str(market_like).split("/")[-1].strip()]:
        c = str(candidate or "").strip()
        if c and c not in candidates:
            candidates.append(c)

    for candidate in candidates:
        res = requests.get(f"{KALSHI_BASE}/markets/{candidate}", timeout=HTTP_TIMEOUT_SEC)
        if not res.ok:
            continue
        payload = res.json()
        market = payload.get("market")
        if isinstance(market, dict):
            return market
    return None


def subcontract_display_name(market):
    """Returns a stable human-readable subcontract name."""
    if not isinstance(market, dict):
        return "Unknown subcontract"

    # In event markets like "Who will win MVP?", the player/candidate is
    # usually stored in yes_sub_title while title is generic across all rows.
    yes_sub = str(market.get("yes_sub_title", "")).strip()
    if yes_sub:
        return yes_sub

    for key in ("title", "subtitle", "ticker"):
        value = str(market.get(key, "")).strip()
        if value:
            return value
    return "Unknown subcontract"


def _normalize_match_text(value):
    chars = []
    for ch in str(value or "").lower():
        if ch.isalnum():
            chars.append(ch)
        else:
            chars.append(" ")
    return " ".join("".join(chars).split())


def _normalize_outcome_label(value):
    return _normalize_match_text(value)


def _extract_outcomes_tokens(poly_market):
    """Return (outcomes, token_ids) aligned from Gamma market payload."""
    if not isinstance(poly_market, dict):
        return [], []
    raw_ids = poly_market.get("clobTokenIds", [])
    if isinstance(raw_ids, str):
        try:
            raw_ids = json.loads(raw_ids)
        except json.JSONDecodeError:
            raw_ids = [raw_ids]
    outcomes = poly_market.get("outcomes") or []
    if isinstance(outcomes, str):
        try:
            outcomes = json.loads(outcomes)
        except json.JSONDecodeError:
            outcomes = [outcomes]
    outcomes = [str(o).strip() for o in outcomes if str(o).strip()]
    token_ids = [str(t).strip() for t in raw_ids if str(t).strip()]
    return outcomes, token_ids


def _match_primary_outcome(outcomes, kalshi_label):
    """Return the outcome label that matches the Kalshi submarket name."""
    norm_k = _normalize_outcome_label(kalshi_label)
    if not outcomes:
        return ""
    # Exact normalized match
    matches = [o for o in outcomes if _normalize_outcome_label(o) == norm_k]
    if len(matches) == 1:
        return matches[0]
    # Fuzzy containment match
    fuzzy = [
        o for o in outcomes
        if norm_k and (norm_k in _normalize_outcome_label(o) or _normalize_outcome_label(o) in norm_k)
    ]
    if len(fuzzy) == 1:
        return fuzzy[0]
    return ""


def _poly_contract_match_text(market):
    if not isinstance(market, dict):
        return ""
    pieces = []
    for key in ("question", "title", "subtitle", "groupItemTitle", "slug"):
        val = str(market.get(key, "")).strip()
        if val:
            pieces.append(val)
    return _normalize_match_text(" ".join(pieces))


# Official Polymarket top-level categories (lowercase for matching)
_POLY_OFFICIAL_CATEGORIES = {
    "crypto", "economics", "mentions", "culture", "weather",
    "finance", "politics", "tech", "sports", "geopolitics",
}


def _extract_category(poly_market: dict) -> str:
    """Return the official Polymarket category from event tags, or 'other'.

    Tags live on the event object. We pull the event slug from market.events[0]
    and do one extra GET /events?slug= to retrieve them, then match against
    Polymarket's official top-level category list.
    """
    events = poly_market.get("events") or []
    event_slug = ""
    if isinstance(events, list) and events and isinstance(events[0], dict):
        event_slug = str(events[0].get("slug") or "").strip()
    if not event_slug:
        return "other"
    try:
        res = requests.get(
            f"{POLY_GAMMA_BASE}/events",
            params={"slug": event_slug},
            timeout=HTTP_TIMEOUT_SEC,
        )
        res.raise_for_status()
        event_data = res.json()
        if isinstance(event_data, list) and event_data:
            event_data = event_data[0]
        tags = event_data.get("tags") or [] if isinstance(event_data, dict) else []
        # First pass: exact match against official categories (lowest ID wins)
        best_label = ""
        best_id = float("inf")
        for tag in tags:
            if not isinstance(tag, dict):
                continue
            label = str(tag.get("label") or tag.get("slug") or "").strip().lower()
            if label not in _POLY_OFFICIAL_CATEGORIES:
                continue
            try:
                tag_id = float(tag.get("id", float("inf")))
            except (ValueError, TypeError):
                tag_id = float("inf")
            if tag_id < best_id:
                best_id = tag_id
                best_label = label
        if best_label:
            return best_label
    except Exception:
        pass
    return "other"


def _build_mapping_row(poly_market, poly_market_slug, kalshi_submarket):
    pair_id = get_next_pair_id()
    # Extract expiry from Polymarket (Gamma API returns endDate or end_date_iso)
    expiry_poly = (
        poly_market.get("endDate")
        or poly_market.get("end_date_iso")
        or poly_market.get("end_date")
        or ""
    )
    # Extract expiry from Kalshi (markets API returns close_time)
    expiry_kalshi = kalshi_submarket.get("close_time") or ""
    outcomes, token_ids = _extract_outcomes_tokens(poly_market)
    outcomes_lc = [o.strip().lower() for o in outcomes]
    primary_outcome = ""
    if len(outcomes_lc) == 2 and "yes" in outcomes_lc and "no" in outcomes_lc:
        primary_outcome = "yes"
    else:
        kalshi_name = str(kalshi_submarket.get("display_name", subcontract_display_name(kalshi_submarket))).strip()
        primary_outcome = _match_primary_outcome(outcomes, kalshi_name)
        if not primary_outcome:
            raise RuntimeError(
                f"Unable to map Kalshi submarket '{kalshi_name}' to Polymarket outcomes {outcomes}"
            )
    event_slug = str(poly_market.get("eventSlug") or poly_market.get("event_slug") or "").strip()
    poly_event_url = f"https://polymarket.com/event/{event_slug}" if event_slug else ""
    _k_ticker = kalshi_submarket["ticker"]
    _k_event_ticker = str(kalshi_submarket.get("event_ticker") or "").strip()
    return {
        "pair_id": pair_id,
        "title_clean": poly_market.get("question", "N/A"),
        "category_tag": _extract_category(poly_market),
        "similarity_score": "1.0",
        "poly_market_id": poly_market.get("conditionId", "N/A"),
        "poly_slug": poly_market_slug,
        "poly_url": f"https://polymarket.com/event/{event_slug}" if event_slug else f"https://polymarket.com/market/{poly_market_slug}",
        "poly_event_url": poly_event_url,
        "poly_outcomes_json": json.dumps(outcomes),
        "poly_token_ids_json": json.dumps(token_ids),
        "poly_primary_outcome": primary_outcome,
        "kalshi_market_id": _k_ticker,
        "kalshi_url": _build_kalshi_url(_k_event_ticker, _k_ticker),
        "expiry_poly_utc": expiry_poly,
        "expiry_kalshi_utc": expiry_kalshi,
    }


def _fetch_markets_for_event(event_ticker: str) -> list:
    """Return all markets for a given event_ticker, paginated."""
    markets = []
    cursor = None
    while True:
        params = {"event_ticker": event_ticker, "limit": 500}
        if cursor:
            params["cursor"] = cursor
        res = requests.get(f"{KALSHI_BASE}/markets", params=params, timeout=HTTP_TIMEOUT_SEC)
        res.raise_for_status()
        payload = res.json()
        page = payload.get("markets", [])
        if isinstance(page, list):
            markets.extend([m for m in page if isinstance(m, dict)])
        cursor = payload.get("cursor")
        if not cursor or not page:
            break
    return markets


def get_submarkets(market_like):
    """Fetches all markets linked to the same Kalshi event.

    Kalshi URL paths can take two forms:
      {event_ticker}/{slug}/{market_ticker}  e.g. kxtesla/tesla-deliveries/kxtesla-26-q1
      {series}/{slug}/{event_ticker}         e.g. kxratecutcount/number-of-rate-cuts/kxratecutcount-26dec31

    We try multiple event_ticker candidates in order and use the first that
    returns at least one market.
    """
    parts = [p for p in str(market_like).split("/") if p]

    # Build ordered candidate list (deduped, uppercased)
    candidates: list[str] = []

    def _add(et: str) -> None:
        et = et.strip().upper()
        if et and et not in candidates:
            candidates.append(et)

    # 1. API lookup of the last segment — if it's a real market ticker the
    #    response includes the authoritative event_ticker.
    if parts:
        market_data = _get_market_payload(parts[-1])
        if isinstance(market_data, dict):
            _add(str(market_data.get("event_ticker", "")))

    # 2. Last URL segment — works when it IS the event ticker (e.g. rate cuts)
    if parts:
        _add(parts[-1])

    # 3. First URL segment — works when it's the base event ticker (e.g. Tesla)
    if len(parts) >= 2:
        _add(parts[0])

    for event_ticker in candidates:
        markets = _fetch_markets_for_event(event_ticker)
        if markets:
            deduped = {}
            for market in markets:
                ticker = str(market.get("ticker", "")).strip()
                if ticker and ticker not in deduped:
                    market["display_name"] = subcontract_display_name(market)
                    deduped[ticker] = market
            return list(deduped.values())

    return []


def _normalize_header_name(value):
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _looks_like_header(first_col, second_col):
    poly_header = _normalize_header_name(first_col)
    kalshi_header = _normalize_header_name(second_col)
    poly_candidates = {
        "polymarket",
        "poly",
        "polymarketlink",
        "polymarketurl",
        "polyurl",
        "polylink",
    }
    kalshi_candidates = {"kalshi", "kalshilink", "kalshiurl"}
    return poly_header in poly_candidates and kalshi_header in kalshi_candidates


def _iter_link_rows(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                yield row
        return
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Reading .xlsx files requires openpyxl. Run: pip install openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                yield list(row)
        finally:
            workbook.close()
        return
    raise RuntimeError("Input file must be .csv or .xlsx")


def load_link_pairs(input_file):
    path = Path(input_file)
    if not path.exists():
        raise RuntimeError(f"Input file does not exist: {input_file}")

    pairs = []
    for row_num, row in enumerate(_iter_link_rows(path), 1):
        first = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        second = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        if not first and not second:
            continue
        if row_num == 1 and _looks_like_header(first, second):
            continue
        if not first or not second:
            raise RuntimeError(
                f"Input row {row_num} must include Polymarket link in column 1 and Kalshi link in column 2"
            )
        pairs.append((row_num, first, second))

    if not pairs:
        raise RuntimeError("Input file has no valid link rows")
    return pairs


def build_output_rows(raw_poly_url, kalshi_url, context_label=""):
    poly_slug = extract_poly_slug(raw_poly_url)
    if not poly_slug:
        raise RuntimeError("Could not extract Polymarket slug from URL")

    poly_contracts = resolve_polymarket_contracts(poly_slug)
    if not poly_contracts:
        raise RuntimeError(f"No Polymarket contracts found for slug: {poly_slug}")

    kalshi_market_like = extract_kalshi_market_like(kalshi_url)
    if not kalshi_market_like:
        raise RuntimeError("Could not extract Kalshi market path from URL")

    submarkets = get_submarkets(kalshi_market_like)
    if not submarkets:
        raise RuntimeError(f"No Kalshi subcontracts found for input: {kalshi_market_like}")

    # Auto-map any Kalshi subcontract where its full normalized name is contained
    # in the normalized Polymarket contract text.
    auto_rows = []
    seen_pairs = set()
    for submarket in submarkets:
        kalshi_name = str(submarket.get("display_name", subcontract_display_name(submarket))).strip()
        kalshi_norm = _normalize_match_text(kalshi_name)
        if not kalshi_norm:
            continue
        for poly_market in poly_contracts:
            poly_text = _poly_contract_match_text(poly_market)
            if kalshi_norm and kalshi_norm in poly_text:
                poly_market_slug = str(poly_market.get("slug", "")).strip() or poly_slug
                pair_key = (str(poly_market.get("conditionId", "")), str(submarket.get("ticker", "")))
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                auto_rows.append(_build_mapping_row(poly_market, poly_market_slug, submarket))

    if auto_rows:
        return auto_rows

    # Fallback behavior: interactive selection when no automatic contains-match found.
    # Loop so the user can pick multiple subcontracts one at a time.
    # Each pick removes that option from the list. Enter '#' or leave blank to stop.
    prefix = f"{context_label} " if context_label else ""
    poly_data = resolve_polymarket_market(poly_slug, context_label=context_label)
    poly_market_slug = str(poly_data.get("slug", "")).strip() or poly_slug

    remaining = list(submarkets)
    manual_rows = []

    while remaining:
        print(f"\n{prefix}Which subcontract do you want to pair? (# or blank to finish)")
        for i, m in enumerate(remaining, 1):
            print(f"{i}) {m.get('display_name', subcontract_display_name(m))}")

        raw = input(f"\n{prefix}Enter choice number: ").strip()
        if not raw or raw == "#":
            break

        try:
            choice = int(raw) - 1
        except ValueError:
            print(f"  Invalid input '{raw}', enter a number or # to stop.")
            continue

        if choice < 0 or choice >= len(remaining):
            print(f"  Choice out of range (1–{len(remaining)}), try again.")
            continue

        selected = remaining.pop(choice)
        manual_rows.append(_build_mapping_row(poly_data, poly_market_slug, selected))

    if not manual_rows:
        raise RuntimeError("No subcontracts selected")

    return manual_rows


# ---------------------------------------------------------------------------
# Live API validation helpers (used by --validate flag)
# ---------------------------------------------------------------------------

def _validate_kalshi_market(ticker: str) -> dict:
    """Verify the Kalshi market is open, not yet closed, and has an accessible orderbook."""
    res = requests.get(f"{KALSHI_BASE}/markets/{ticker}", timeout=HTTP_TIMEOUT_SEC)
    res.raise_for_status()
    market = res.json().get("market", {})
    if not isinstance(market, dict):
        raise RuntimeError(f"Kalshi market payload malformed for {ticker}")
    status = str(market.get("status", "")).strip().lower()
    if status not in {"open", "active"}:
        raise RuntimeError(f"Kalshi market '{ticker}' status is '{status}', expected active/open")

    close_time = str(market.get("close_time") or "").strip()
    if close_time:
        try:
            close_dt = datetime.datetime.fromisoformat(close_time.rstrip("Z")).replace(
                tzinfo=datetime.timezone.utc
            )
            if close_dt <= datetime.datetime.now(datetime.timezone.utc):
                raise RuntimeError(
                    f"Kalshi market '{ticker}' has already closed (close_time={close_time})"
                )
        except ValueError:
            pass  # Unparseable date — skip the check

    orderbook_res = requests.get(
        f"{KALSHI_BASE}/markets/{ticker}/orderbook", timeout=HTTP_TIMEOUT_SEC
    )
    orderbook_res.raise_for_status()

    yes_sub = str(market.get("yes_sub_title", "")).strip()
    display_name = (
        yes_sub
        or str(market.get("title", "")).strip()
        or str(market.get("subtitle", "")).strip()
    )
    return {"close_time": close_time, "display_name": display_name}


def _validate_poly_clob(slug: str) -> dict:
    """Verify Polymarket CLOB tokens are accessible and prices look sane."""
    res = requests.get(f"{POLY_GAMMA_BASE}/markets", params={"slug": slug}, timeout=HTTP_TIMEOUT_SEC)
    res.raise_for_status()
    payload = res.json()
    if isinstance(payload, list) and payload:
        market = payload[0]
    elif isinstance(payload, dict) and payload.get("conditionId"):
        market = payload
    else:
        raise RuntimeError(f"No Polymarket market found for slug '{slug}'")

    outcomes, token_ids = _extract_outcomes_tokens(market)
    if len(token_ids) < 2:
        raise RuntimeError(f"Polymarket market '{slug}' does not expose at least two CLOB token IDs")

    outcomes_lc = [o.strip().lower() for o in outcomes]
    if "yes" in outcomes_lc and "no" in outcomes_lc:
        yes_idx = outcomes_lc.index("yes")
        no_idx = outcomes_lc.index("no")
        if token_ids[yes_idx] == token_ids[no_idx]:
            raise RuntimeError(f"Polymarket market '{slug}' YES and NO token IDs are identical")

    for token_id in token_ids:
        book_res = requests.get(
            f"{POLY_CLOB_HOST}/book", params={"token_id": token_id}, timeout=HTTP_TIMEOUT_SEC
        )
        book_res.raise_for_status()

    yes_ask = float(market.get("bestAsk") or 0)
    yes_bid = float(market.get("bestBid") or 0)
    if yes_ask > 0 and yes_bid > 0 and "yes" in outcomes_lc and "no" in outcomes_lc:
        no_ask = 1.0 - yes_bid
        ask_sum = yes_ask + no_ask
        if ask_sum > 1.10:
            print(
                f"  ⚠ Polymarket '{slug}' ask prices sum to {ask_sum:.3f} "
                f"(YES ask={yes_ask:.3f}, NO ask={no_ask:.3f}) — allowing anyway"
            )

    end_date = str(
        market.get("endDate") or market.get("end_date_iso") or market.get("end_date") or ""
    ).strip()
    return {"end_date": end_date}


def validate_row(row: dict) -> dict:
    """Run live API validation on a generated row.  Raises RuntimeError on failure.

    Checks:
    - Kalshi market is open/active and not yet closed
    - Kalshi orderbook endpoint is accessible
    - Polymarket CLOB token endpoints are accessible
    - Ask-price sanity (sum > 1.10 logs a warning but no longer rejects)
    Also populates resolution_time_utc if not already set.
    """
    kalshi_ticker = str(row.get("kalshi_market_id", "")).strip()
    poly_slug = str(row.get("poly_slug", "")).strip()
    if not kalshi_ticker or not poly_slug:
        raise RuntimeError("Row missing kalshi_market_id or poly_slug")

    kalshi_info = _validate_kalshi_market(kalshi_ticker)
    poly_info = _validate_poly_clob(poly_slug)

    validated = dict(row)
    if not validated.get("resolution_time_utc"):
        validated["resolution_time_utc"] = (
            kalshi_info.get("close_time") or poly_info.get("end_date") or ""
        )
    return validated


def parse_args():
    parser = argparse.ArgumentParser(
        description="Create mapping rows from Polymarket and Kalshi links (interactive or file batch mode)."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        help="Optional .csv or .xlsx file where column 1 is Polymarket link and column 2 is Kalshi link.",
    )
    parser.add_argument(
        "--user",
        default=None,
        help="Username for per-user pairs subfolder (e.g. --user dplynn). "
             "Pairs are written to Polytoken/pairs/{username}/pairs.csv. "
             "Defaults to the OS login name if omitted.",
    )
    parser.add_argument(
        "--no-validate",
        action="store_true",
        help="Skip live API validation (Kalshi market open, CLOB token accessibility, price sanity). "
             "Validation runs by default.",
    )
    return parser.parse_args()


def _is_duplicate(row, existing_keys):
    """Check if a row's (poly_market_id, kalshi_market_id) already exists."""
    pair_key = (
        str(row.get("poly_market_id", "")).strip(),
        str(row.get("kalshi_market_id", "")).strip(),
    )
    if pair_key[0] and pair_key[1] and pair_key in existing_keys:
        return True
    return False


def _open_user_writer(username: str):
    """Open the per-user pairs CSV for appending.  Returns (file_handle, csv.DictWriter)."""
    out_dir = USER_PAIRS_DIR / username
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "pairs.csv"
    needs_header = not out_path.exists() or out_path.stat().st_size == 0
    fh = out_path.open("a", newline="", encoding="utf-8")
    writer = csv.DictWriter(fh, fieldnames=OUTPUT_FIELDS)
    if needs_header:
        writer.writeheader()
    return fh, writer, out_path


def main():
    args = parse_args()

    # Resolve username — falls back to OS login name automatically
    username = (args.user or getpass.getuser()).strip().lower()
    if not username:
        print("ERROR: could not determine username; pass --user explicitly", file=sys.stderr)
        sys.exit(1)

    existing_keys = load_existing_pair_keys()
    if existing_keys:
        print(f"Loaded {len(existing_keys)} existing pairs for duplicate checking", file=sys.stderr)

    duplicates_skipped = 0
    fh, writer, out_path = _open_user_writer(username)

    def _maybe_validate(row, label=""):
        if args.no_validate:
            return row
        try:
            return validate_row(row)
        except Exception as exc:
            print(f"{label}Validation failed for {row.get('kalshi_market_id', '?')}: {exc}", file=sys.stderr)
            return None

    try:
        if args.input_file:
            pairs = load_link_pairs(args.input_file)
            successes = 0
            for row_num, poly_link, kalshi_link in pairs:
                context_label = f"[row {row_num}] "
                try:
                    rows = build_output_rows(poly_link, kalshi_link, context_label=f"[row {row_num}]")
                except Exception as exc:
                    print(f"[row {row_num}] Failed: {exc}", file=sys.stderr)
                    continue
                for row in rows:
                    row = _maybe_validate(row, label=context_label)
                    if row is None:
                        continue
                    if _is_duplicate(row, existing_keys):
                        print(f"{context_label}Skipped duplicate: {row.get('kalshi_market_id', '?')}", file=sys.stderr)
                        duplicates_skipped += 1
                        continue
                    writer.writerow(row)
                    successes += 1

            if duplicates_skipped:
                print(f"Skipped {duplicates_skipped} duplicate pair(s)", file=sys.stderr)
            if successes == 0 and duplicates_skipped == 0:
                raise RuntimeError("No rows were processed successfully")
        else:
            raw_poly_url = input("Paste Polymarket link: ").strip()
            kalshi_url = input("Paste Kalshi link: ").strip()
            rows = build_output_rows(raw_poly_url, kalshi_url)
            for row in rows:
                row = _maybe_validate(row)
                if row is None:
                    continue
                if _is_duplicate(row, existing_keys):
                    print(f"Skipped duplicate: {row.get('kalshi_market_id', '?')}", file=sys.stderr)
                    duplicates_skipped += 1
                    continue
                writer.writerow(row)

            if duplicates_skipped:
                print(f"Skipped {duplicates_skipped} duplicate pair(s)", file=sys.stderr)
    finally:
        fh.close()

    print(f"Pairs written to {out_path}", file=sys.stderr)

if __name__ == "__main__":
    main()
