import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

import requests

try:
    from polytoken import (
        HTTP_TIMEOUT_SEC,
        KALSHI_BASE,
        OUTPUT_FIELDS,
        POLY_GAMMA_BASE,
        build_output_rows,
        load_link_pairs,
    )
except ModuleNotFoundError:
    from Polytoken.polytoken import (
        HTTP_TIMEOUT_SEC,
        KALSHI_BASE,
        OUTPUT_FIELDS,
        POLY_GAMMA_BASE,
        build_output_rows,
        load_link_pairs,
    )


V2_CONTRACTS_FILENAME = "Pairs_for_Kalshi_and_Polymarket.xlsx"


def _resolve_default_v2_contracts_path() -> Path:
    script_path = Path(__file__).resolve()
    for base in script_path.parents:
        candidate = base / "V5" / V2_CONTRACTS_FILENAME
        if candidate.exists():
            return candidate
    # Fallback to normal repo layout even if the file doesn't exist yet.
    return script_path.parents[1] / "V5" / V2_CONTRACTS_FILENAME


DEFAULT_V2_CONTRACTS_PATH = _resolve_default_v2_contracts_path()
POLY_CLOB_HOST = "https://clob.polymarket.com"


def _normalize_header_name(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _parse_list_field(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(v) for v in value]
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(v) for v in parsed]
        except json.JSONDecodeError:
            return [text]
    return []


def _normalize_outcome_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    cleaned = []
    for ch in text:
        cleaned.append(ch if ch.isalnum() else " ")
    return " ".join("".join(cleaned).split())


def _match_primary_outcome(outcomes: list[str], kalshi_label: str) -> str:
    norm_k = _normalize_outcome_label(kalshi_label)
    if not outcomes:
        return ""
    matches = [o for o in outcomes if _normalize_outcome_label(o) == norm_k]
    if len(matches) == 1:
        return matches[0]
    fuzzy = [
        o for o in outcomes
        if norm_k and (norm_k in _normalize_outcome_label(o) or _normalize_outcome_label(o) in norm_k)
    ]
    if len(fuzzy) == 1:
        return fuzzy[0]
    return ""


def _validate_kalshi_market(ticker: str) -> dict[str, Any]:
    res = requests.get(f"{KALSHI_BASE}/markets/{ticker}", timeout=HTTP_TIMEOUT_SEC)
    res.raise_for_status()
    market = res.json().get("market", {})
    if not isinstance(market, dict):
        raise RuntimeError(f"Kalshi market payload malformed for {ticker}")
    status = str(market.get("status", "")).strip().lower()
    if status not in {"open", "active"}:
        raise RuntimeError(f"Kalshi market '{ticker}' status is '{status}', expected active/open")

    # Verify the market has not yet closed
    import datetime as _dt
    close_time = str(market.get("close_time") or "").strip()
    if close_time:
        try:
            close_dt = _dt.datetime.fromisoformat(close_time.rstrip("Z")).replace(tzinfo=_dt.timezone.utc)
            if close_dt <= _dt.datetime.now(_dt.timezone.utc):
                raise RuntimeError(f"Kalshi market '{ticker}' has already closed (close_time={close_time})")
        except ValueError:
            pass  # Unparseable date — skip the check

    orderbook_res = requests.get(f"{KALSHI_BASE}/markets/{ticker}/orderbook", timeout=HTTP_TIMEOUT_SEC)
    orderbook_res.raise_for_status()

    yes_sub = str(market.get("yes_sub_title", "")).strip()
    display_name = yes_sub or str(market.get("title", "")).strip() or str(market.get("subtitle", "")).strip()

    return {"close_time": close_time, "display_name": display_name}


def _resolve_poly_market(slug: str) -> dict[str, Any]:
    res = requests.get(f"{POLY_GAMMA_BASE}/markets", params={"slug": slug}, timeout=HTTP_TIMEOUT_SEC)
    res.raise_for_status()
    payload = res.json()
    if isinstance(payload, list) and payload:
        return payload[0]
    if isinstance(payload, dict) and payload.get("conditionId"):
        return payload
    raise RuntimeError(f"No Polymarket market found for slug '{slug}'")


def _validate_poly_market(slug: str) -> dict[str, Any]:
    market = _resolve_poly_market(slug)
    token_ids = _parse_list_field(market.get("clobTokenIds"))
    outcomes_raw = _parse_list_field(market.get("outcomes"))
    outcomes = [str(o).strip() for o in outcomes_raw if str(o).strip()]
    outcomes_lc = [o.strip().lower() for o in outcomes]
    if len(token_ids) < 2:
        raise RuntimeError(f"Polymarket market '{slug}' does not expose at least two CLOB token IDs")

    if "yes" in outcomes_lc and "no" in outcomes_lc:
        yes_idx = outcomes_lc.index("yes")
        no_idx = outcomes_lc.index("no")
        yes_id = token_ids[yes_idx]
        no_id = token_ids[no_idx]
        if yes_id == no_id:
            raise RuntimeError(
                f"Polymarket market '{slug}' YES and NO token IDs are identical ({yes_id!r})"
            )

    # Verify CLOB tokens are accessible
    for token_id in token_ids:
        book_res = requests.get(f"{POLY_CLOB_HOST}/book", params={"token_id": token_id}, timeout=HTTP_TIMEOUT_SEC)
        book_res.raise_for_status()

    # Sanity check using Gamma API prices (only for yes/no)
    yes_ask = float(market.get("bestAsk") or 0)
    yes_bid = float(market.get("bestBid") or 0)
    if yes_ask > 0 and yes_bid > 0 and "yes" in outcomes_lc and "no" in outcomes_lc:
        no_ask = 1.0 - yes_bid
        ask_sum = yes_ask + no_ask
        if ask_sum > 1.10:
            raise RuntimeError(
                f"Polymarket '{slug}' ask prices sum to {ask_sum:.3f} "
                f"(YES ask={yes_ask:.3f}, NO ask?{no_ask:.3f})"
            )

    end_date = str(
        market.get("endDate") or market.get("end_date_iso") or market.get("end_date") or ""
    ).strip()

    return {
        "conditionId": str(market.get("conditionId", "")).strip(),
        "slug": str(market.get("slug", "")).strip() or slug,
        "end_date": end_date,
        "outcomes": outcomes,
        "token_ids": token_ids,
        "event_slug": str(market.get("eventSlug") or market.get("event_slug") or "").strip(),
    }


def validate_row(row: dict[str, str]) -> dict[str, str]:
    kalshi_market_id = str(row.get("kalshi_market_id", "")).strip()
    poly_slug = str(row.get("poly_slug", "")).strip()
    if not kalshi_market_id or not poly_slug:
        raise RuntimeError("Row missing kalshi_market_id or poly_slug")

    kalshi_info = _validate_kalshi_market(kalshi_market_id)
    poly = _validate_poly_market(poly_slug)

    validated = dict(row)
    if poly.get("conditionId"):
        validated["poly_market_id"] = poly["conditionId"]
    if poly.get("slug"):
        validated["poly_slug"] = poly["slug"]
        validated["poly_url"] = f"https://polymarket.com/market/{poly['slug']}"

    outcomes = poly.get("outcomes", []) or []
    token_ids = poly.get("token_ids", []) or []
    outcomes_lc = [str(o).strip().lower() for o in outcomes]
    primary_outcome = ""
    if len(outcomes_lc) == 2 and "yes" in outcomes_lc and "no" in outcomes_lc:
        primary_outcome = "yes"
    else:
        kalshi_label = kalshi_info.get("display_name", "")
        primary_outcome = _match_primary_outcome(outcomes, kalshi_label)
        if not primary_outcome:
            raise RuntimeError(
                f"Unable to map Kalshi submarket '{kalshi_label}' to Polymarket outcomes {outcomes}"
            )
    validated["poly_outcomes_json"] = json.dumps(outcomes)
    validated["poly_token_ids_json"] = json.dumps(token_ids)
    validated["poly_primary_outcome"] = primary_outcome
    if poly.get("event_slug"):
        validated["poly_event_url"] = f"https://polymarket.com/event/{poly['event_slug']}"

    # Populate resolution_time_utc: prefer Kalshi close_time, fall back to Poly end_date
    if not validated.get("resolution_time_utc"):
        validated["resolution_time_utc"] = (
            kalshi_info.get("close_time") or poly.get("end_date") or ""
        )

    return validated


def _existing_pair_keys(ws, header_map: dict[str, int]) -> set[tuple[str, str]]:
    poly_col = header_map.get(_normalize_header_name("poly_market_id"))
    kalshi_col = header_map.get(_normalize_header_name("kalshi_market_id"))
    keys: set[tuple[str, str]] = set()
    if not poly_col or not kalshi_col:
        return keys

    for r in range(2, ws.max_row + 1):
        poly_id = str(ws.cell(row=r, column=poly_col).value or "").strip()
        kalshi_id = str(ws.cell(row=r, column=kalshi_col).value or "").strip()
        if poly_id and kalshi_id:
            keys.add((poly_id, kalshi_id))
    return keys


def append_rows_to_contract_list(rows: list[dict[str, str]], target_path: Path) -> tuple[int, int]:
    if target_path.suffix.lower() != ".xlsx":
        raise RuntimeError(f"Only .xlsx target is supported for direct append, got: {target_path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Appending to .xlsx requires openpyxl. Run: pip install openpyxl") from exc

    if not target_path.exists():
        raise RuntimeError(f"Target contracts file not found: {target_path}")

    wb = load_workbook(target_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {_normalize_header_name(h): i + 1 for i, h in enumerate(headers) if _normalize_header_name(h)}
    if not header_map:
        wb.close()
        raise RuntimeError("Target sheet header row is empty")

    # Ensure new output fields exist in the header row
    for field in OUTPUT_FIELDS:
        key = _normalize_header_name(field)
        if key and key not in header_map:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=field)
            header_map[key] = new_col

    existing_keys = _existing_pair_keys(ws, header_map)
    inserted = 0
    skipped = 0

    for row in rows:
        pair_key = (
            str(row.get("poly_market_id", "")).strip(),
            str(row.get("kalshi_market_id", "")).strip(),
        )
        if pair_key[0] and pair_key[1] and pair_key in existing_keys:
            skipped += 1
            continue

        new_row_index = ws.max_row + 1
        for field in OUTPUT_FIELDS:
            col = header_map.get(_normalize_header_name(field))
            if col:
                ws.cell(row=new_row_index, column=col, value=row.get(field, ""))

        # Write resolution_time_utc if the column exists
        res_col = header_map.get(_normalize_header_name("resolution_time_utc"))
        if res_col and row.get("resolution_time_utc"):
            ws.cell(row=new_row_index, column=res_col, value=row["resolution_time_utc"])

        active_col = header_map.get(_normalize_header_name("active"))
        if active_col:
            ws.cell(row=new_row_index, column=active_col, value=True)

        if pair_key[0] and pair_key[1]:
            existing_keys.add(pair_key)
        inserted += 1

    wb.save(target_path)
    wb.close()
    return inserted, skipped


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Generate mapping rows from Polymarket+Kalshi links, validate API pullability, "
            "append valid rows into V2 contract list, then print valid rows to stdout."
        )
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        help="Optional .csv or .xlsx file where column 1 is Polymarket link and column 2 is Kalshi link.",
    )
    parser.add_argument(
        "--target",
        default=str(DEFAULT_V2_CONTRACTS_PATH),
        help="Path to V2 contracts .xlsx file to append to.",
    )
    return parser.parse_args()


def _collect_candidate_rows(args) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []
    if args.input_file:
        pairs = load_link_pairs(args.input_file)
        for row_num, poly_link, kalshi_link in pairs:
            context_label = f"[row {row_num}]"
            try:
                generated = build_output_rows(poly_link, kalshi_link, context_label=context_label)
                candidates.extend(generated)
            except Exception as exc:
                print(f"{context_label} Failed to generate rows: {exc}", file=sys.stderr)
        return candidates

    raw_poly_url = input("Paste Polymarket link: ").strip()
    kalshi_url = input("Paste Kalshi link: ").strip()
    return build_output_rows(raw_poly_url, kalshi_url)


def main():
    args = parse_args()
    target = Path(args.target).resolve()
    candidates = _collect_candidate_rows(args)
    if not candidates:
        raise RuntimeError("No candidate rows were generated")

    valid_rows: list[dict[str, str]] = []
    for row in candidates:
        pair_id = str(row.get("pair_id", "unknown")).strip() or "unknown"
        try:
            valid_rows.append(validate_row(row))
        except Exception as exc:
            print(f"[{pair_id}] Validation failed: {exc}", file=sys.stderr)

    if not valid_rows:
        raise RuntimeError("No rows passed API validation")

    inserted, skipped = append_rows_to_contract_list(valid_rows, target)
    print(f"Appended {inserted} row(s) to {target}; skipped {skipped} duplicate row(s).", file=sys.stderr)

    writer = csv.DictWriter(sys.stdout, fieldnames=OUTPUT_FIELDS)
    for row in valid_rows:
        writer.writerow(row)


if __name__ == "__main__":
    main()
