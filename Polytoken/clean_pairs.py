"""clean_pairs.py — Remove expired and duplicate rows from the V5 pairs Excel file.

Usage:
    python clean_pairs.py                          # dry-run (show what would be removed)
    python clean_pairs.py --apply                  # actually remove rows and save
    python clean_pairs.py --target /path/to/file   # explicit pairs file path
    python clean_pairs.py --apply --keep-active    # only remove rows where active=False/expired

Expired:  resolution_time_utc is in the past (< today).
Duplicate: same (poly_market_id, kalshi_market_id) key appears more than once; keeps
           the first occurrence.
"""

from __future__ import annotations

import argparse
import datetime
import sys
from pathlib import Path
from typing import Any


# ── Default path resolution (same logic as check_and_append_v2.py) ────────────

V2_CONTRACTS_FILENAME = "Pairs_for_Kalshi_and_Polymarket.xlsx"


def _resolve_default_path() -> Path:
    script_path = Path(__file__).resolve()
    for base in script_path.parents:
        candidate = base / "V5" / V2_CONTRACTS_FILENAME
        if candidate.exists():
            return candidate
    return script_path.parents[1] / "V5" / V2_CONTRACTS_FILENAME


DEFAULT_PATH = _resolve_default_path()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _parse_date(value: Any) -> datetime.date | None:
    """Try to parse a cell value as a date. Returns None if unparseable."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # Try ISO date prefix
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            return datetime.datetime.strptime(s[:19], fmt[:len(s[:19].replace("T", " ").split()[0])]).date()
        except ValueError:
            pass
    # Fallback: just grab the first 10 chars
    try:
        return datetime.date.fromisoformat(s[:10])
    except ValueError:
        return None


def _cell_str(value: Any) -> str:
    return str(value or "").strip()


# ── Main clean logic ──────────────────────────────────────────────────────────

def clean_pairs(target_path: Path, apply: bool) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not target_path.exists():
        print(f"ERROR: pairs file not found: {target_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(target_path)
    ws = wb.active
    today = datetime.date.today()

    # Build header map
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map: dict[str, int] = {
        _normalize_header(h): i + 1
        for i, h in enumerate(headers)
        if _normalize_header(h)
    }

    # Locate key columns
    resolution_col = (
        header_map.get(_normalize_header("resolution_time_utc"))
        or header_map.get(_normalize_header("resolution_date"))
    )
    poly_id_col = header_map.get(_normalize_header("poly_market_id"))
    kalshi_id_col = header_map.get(_normalize_header("kalshi_market_id"))
    active_col = header_map.get(_normalize_header("active"))
    pair_id_col = header_map.get(_normalize_header("pair_id"))

    print(f"File: {target_path}")
    print(f"Today: {today}")
    print(f"Columns found: resolution={resolution_col}, poly_id={poly_id_col}, "
          f"kalshi_id={kalshi_id_col}, active={active_col}, pair_id={pair_id_col}")
    print()

    # Scan rows (data starts at row 2)
    max_row = ws.max_row
    rows_to_delete: list[int] = []
    seen_pair_keys: set[tuple[str, str]] = set()
    expired_count = 0
    duplicate_count = 0

    for r in range(2, max_row + 1):
        pair_id = _cell_str(ws.cell(row=r, column=pair_id_col).value) if pair_id_col else f"row {r}"

        # Skip entirely blank rows
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        reason = None

        # Expired check
        if resolution_col:
            res_val = ws.cell(row=r, column=resolution_col).value
            res_date = _parse_date(res_val)
            if res_date is not None and res_date < today:
                reason = f"expired (resolution={res_date})"
                expired_count += 1

        # Duplicate check (only if not already expired)
        if reason is None and poly_id_col and kalshi_id_col:
            poly_id = _cell_str(ws.cell(row=r, column=poly_id_col).value)
            kalshi_id = _cell_str(ws.cell(row=r, column=kalshi_id_col).value)
            if poly_id and kalshi_id:
                key = (poly_id, kalshi_id)
                if key in seen_pair_keys:
                    reason = f"duplicate of ({poly_id[:12]}..., {kalshi_id})"
                    duplicate_count += 1
                else:
                    seen_pair_keys.add(key)

        if reason:
            rows_to_delete.append(r)
            action = "REMOVE" if apply else "would remove"
            print(f"  [{action}] row {r:>4} | pair_id={pair_id:<30} | {reason}")

    print()
    print(f"Summary: {expired_count} expired, {duplicate_count} duplicates → "
          f"{len(rows_to_delete)} row(s) to remove out of {max_row - 1} data rows")

    if not rows_to_delete:
        print("Nothing to clean.")
        wb.close()
        return

    if not apply:
        print("\nDry-run: pass --apply to actually remove these rows.")
        wb.close()
        return

    # Delete rows in reverse order to preserve row indices
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    wb.save(target_path)
    wb.close()
    print(f"\nSaved {target_path} — removed {len(rows_to_delete)} row(s).")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Remove expired and duplicate rows from the V5 pairs Excel file."
    )
    parser.add_argument(
        "--target",
        default=str(DEFAULT_PATH),
        help=f"Path to the pairs .xlsx file (default: {DEFAULT_PATH})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually delete rows and save. Without this flag the script runs in dry-run mode.",
    )
    args = parser.parse_args()
    clean_pairs(Path(args.target).resolve(), apply=args.apply)


if __name__ == "__main__":
    main()
