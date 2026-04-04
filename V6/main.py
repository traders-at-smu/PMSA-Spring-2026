"""V4 CLI entry point — Kalshi × Polymarket cross-market arbitrage bot.

Usage
-----
    python main.py [--config config.json] COMMAND

Commands:
    validate   Check config file and pairs list, then exit.
    scan       Run one scan cycle, print results, then exit (no trades placed).
    run        Scan and trade continuously until Ctrl+C.

Configuration is loaded from config.json (or config.example.json as fallback).
Lines starting with // are treated as comments and ignored.

Paper mode is the default. Live trading requires API credentials in config.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

from colorama import Fore, Style, init as colorama_init

from connectors import KalshiConnector, PolymarketConnector, load_pairs
from fees import parse_formula, validate_formula
from bot import run_scan, run_loop, _resolve_poly_fee_rate

_SCRIPT_DIR = Path(__file__).parent


# ── Config loading ────────────────────────────────────────────────────────────

def _strip_comments(text: str) -> str:
    """Remove lines whose first non-whitespace characters are //."""
    return "\n".join(
        line for line in text.splitlines()
        if not line.strip().startswith("//")
    )


def load_config(path: str) -> dict:
    p = Path(path)
    if not p.exists():
        # Fall back to the bundled example config
        example = _SCRIPT_DIR / "config.example.json"
        if example.exists():
            p = example
        else:
            print(f"{Fore.RED}Config not found:{Style.RESET_ALL} {path}", file=sys.stderr)
            sys.exit(1)

    try:
        text = p.read_text(encoding="utf-8")
        cfg = json.loads(_strip_comments(text))
    except json.JSONDecodeError as exc:
        print(f"{Fore.RED}Config parse error in {p}:{Style.RESET_ALL} {exc}", file=sys.stderr)
        sys.exit(1)

    # Remove documentation keys (any key starting with _) at all levels
    _strip_doc_keys(cfg)
    return cfg


def _strip_doc_keys(obj) -> None:
    """Recursively remove keys starting with '_' (used for inline documentation)."""
    if isinstance(obj, dict):
        for key in [k for k in obj if k.startswith("_")]:
            del obj[key]
        for v in obj.values():
            _strip_doc_keys(v)


def _validate_config(cfg: dict) -> list[str]:
    """Return a list of error strings (empty list = config is valid)."""
    errors: list[str] = []

    for key in ("mode", "min_arr", "max_contracts", "fees", "pairs_file"):
        if key not in cfg:
            errors.append(f"Missing required key: '{key}'")

    if "mode" in cfg and cfg["mode"] not in ("paper", "live"):
        errors.append("'mode' must be 'paper' or 'live'")

    if "fees" in cfg:
        for venue in ("kalshi", "polymarket"):
            if venue not in cfg["fees"]:
                errors.append(f"Missing fees.{venue} section")
            elif "formula" not in cfg["fees"][venue]:
                errors.append(f"Missing fees.{venue}.formula")
            else:
                try:
                    validate_formula(cfg["fees"][venue]["formula"], f"fees.{venue}.formula")
                except ValueError as exc:
                    errors.append(str(exc))

    return errors


def _compile_fee_fns(cfg: dict) -> None:
    """Parse fee formula strings into callables and store them in cfg."""
    for venue in ("kalshi", "polymarket"):
        cfg["fees"][venue]["_fn"] = parse_formula(cfg["fees"][venue]["formula"])


def _effective_mode(cfg: dict) -> str:
    """Return 'live' only if mode=live AND both sets of API keys are present."""
    if cfg.get("mode") != "live":
        return "paper"
    k = cfg.get("kalshi", {})
    p = cfg.get("polymarket", {})
    has_kalshi = bool(k.get("api_key")) and bool(k.get("private_key_base64"))
    has_poly = bool(p.get("private_key"))
    return "live" if (has_kalshi and has_poly) else "paper"


# ── Banner ────────────────────────────────────────────────────────────────────

def _print_banner(cfg: dict, pairs_count: int) -> None:
    art_path = _SCRIPT_DIR / "ascii-art(1).txt"
    if art_path.exists():
        art = art_path.read_text(encoding="utf-8", errors="replace")
        print(Fore.CYAN + art.rstrip() + Style.RESET_ALL)

    mode = cfg["mode"]
    mode_str = (
        f"{Fore.GREEN}{Style.BRIGHT}LIVE{Style.RESET_ALL}"
        if mode == "live"
        else f"{Fore.YELLOW}{Style.BRIGHT}PAPER{Style.RESET_ALL}"
    )
    arr_pct = float(cfg.get("min_arr", 0.1)) * 100
    max_contracts = int(cfg.get("max_contracts", 0))
    log_path = cfg.get("trade_log", "trades.json")
    interval = cfg.get("scan_interval_seconds", 5)

    label_w = 12
    print(f"\n  {Style.BRIGHT}Kalshi × Polymarket Arb Bot{Style.RESET_ALL}  v5.0")
    print(f"  {'─' * 42}")
    print(f"  {'Mode':<{label_w}}{mode_str}")
    print(f"  {'Min ARR':<{label_w}}{arr_pct:.1f}%")
    print(f"  {'Max contracts':<{label_w}}{max_contracts} per trade")
    print(f"  {'Interval':<{label_w}}{interval}s")
    print(f"  {'Pairs':<{label_w}}{pairs_count} active")
    print(f"  {'Trade log':<{label_w}}{log_path}")
    print()


# ── Connector factory ─────────────────────────────────────────────────────────

def _build_connectors(cfg: dict) -> tuple[KalshiConnector, PolymarketConnector]:
    k = cfg.get("kalshi", {})
    p = cfg.get("polymarket", {})
    api = cfg.get("api", {})
    kalshi = KalshiConnector(
        api_key=k.get("api_key", ""),
        private_key_base64=k.get("private_key_base64", ""),
        base_url=api.get("kalshi_base_url", ""),
    )
    poly = PolymarketConnector(
        private_key=p.get("private_key", ""),
        clob_url=api.get("polymarket_clob_url", ""),
        gamma_url=api.get("polymarket_gamma_url", ""),
        api_key=p.get("api_key", ""),
        api_secret=p.get("api_secret", ""),
        api_passphrase=p.get("api_passphrase", ""),
    )
    return kalshi, poly


# ── Commands ──────────────────────────────────────────────────────────────────

def cmd_validate(args) -> int:
    cfg = load_config(args.config)
    errors = _validate_config(cfg)

    if errors:
        for e in errors:
            print(f"  {Fore.RED}✗{Style.RESET_ALL} {e}")
        return 1

    pairs_file = cfg.get("pairs_file", "")
    if not Path(pairs_file).exists():
        print(f"  {Fore.RED}✗{Style.RESET_ALL} Pairs file not found: {pairs_file}")
        return 1

    pairs = load_pairs(pairs_file)
    mode = _effective_mode(cfg)

    print(f"  {Fore.GREEN}✓{Style.RESET_ALL} Config valid")
    print(f"  {Fore.GREEN}✓{Style.RESET_ALL} Pairs file: {len(pairs)} active pair(s) in '{pairs_file}'")

    if cfg.get("mode") == "live" and mode == "paper":
        print(
            f"  {Fore.YELLOW}!{Style.RESET_ALL} mode=live but API keys are missing "
            "— will run in PAPER mode"
        )
    else:
        print(f"  {Fore.GREEN}✓{Style.RESET_ALL} Effective mode: {mode.upper()}")

    print(f"\n  Fee formulas:")
    for venue in ("kalshi", "polymarket"):
        formula = cfg.get("fees", {}).get(venue, {}).get("formula", "(not set)")
        print(f"    {venue:<12}{formula}")

    return 0


def _merge_user_pairs(pairs_file: str, user_pairs_dirs: str | list | None) -> None:
    """Merge per-user staging CSVs into the master pairs Excel at startup.

    Accepts a single directory path or a list of directory paths.  Each
    directory is scanned for ``{username}/pairs.csv`` files.  New rows are
    appended to the master Excel (deduped by poly_market_id + kalshi_market_id)
    and each CSV is cleared after a successful write.
    """
    if not user_pairs_dirs:
        return

    # Normalise to a list of resolved Paths
    raw = [user_pairs_dirs] if isinstance(user_pairs_dirs, str) else list(user_pairs_dirs)
    dirs: list[Path] = []
    pairs_parent = Path(pairs_file).parent
    for entry in raw:
        p = Path(entry)
        if not p.is_absolute():
            p = pairs_parent / entry
        if p.exists():
            dirs.append(p)

    if not dirs:
        return

    csv_files = sorted(f for d in dirs for f in d.glob("*/pairs.csv"))
    if not csv_files:
        return

    master = Path(pairs_file)
    if not master.exists() or not master.suffix.lower() == ".xlsx":
        print(
            f"  [merge] pairs_file is not an .xlsx — skipping user pairs merge",
            file=sys.stderr,
        )
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [merge] openpyxl not installed — skipping user pairs merge", file=sys.stderr)
        return

    # Read the master Excel to build the existing-keys set
    def _load_master_keys(ws) -> set[tuple[str, str]]:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hmap = {str(h or "").strip().lower(): i + 1 for i, h in enumerate(headers)}
        poly_col = hmap.get("poly_market_id")
        kalshi_col = hmap.get("kalshi_market_id")
        keys: set[tuple[str, str]] = set()
        if poly_col is None or kalshi_col is None:
            return keys
        for r in range(2, ws.max_row + 1):
            pid = str(ws.cell(row=r, column=poly_col).value or "").strip()
            kid = str(ws.cell(row=r, column=kalshi_col).value or "").strip()
            if pid and kid:
                keys.add((pid, kid))
        return keys

    def _col_map(ws) -> dict[str, int]:
        """Return {field_name: column_index} for the master sheet header row."""
        return {
            str(ws.cell(row=1, column=c).value or "").strip(): c
            for c in range(1, ws.max_column + 1)
        }

    total_added = 0
    total_skipped = 0
    for csv_file in csv_files:
        try:
            with csv_file.open(newline="", encoding="utf-8") as f:
                rows = list(csv.DictReader(f))
            if not rows:
                continue

            wb = load_workbook(master)
            ws = wb.active
            existing = _load_master_keys(ws)
            col_map = _col_map(ws)
            added = 0
            skipped = 0

            for row in rows:
                pid = str(row.get("poly_market_id") or "").strip()
                kid = str(row.get("kalshi_market_id") or "").strip()
                if not pid or not kid or (pid, kid) in existing:
                    skipped += 1
                    continue
                next_r = ws.max_row + 1
                for field, value in row.items():
                    col = col_map.get(field)
                    if col:
                        ws.cell(row=next_r, column=col, value=value)
                existing.add((pid, kid))
                added += 1

            if added:
                wb.save(master)
            total_added += added
            total_skipped += skipped

            # Clear the CSV after successful write (keep file, remove content)
            csv_file.write_text("", encoding="utf-8")
            print(
                f"  [merge] {csv_file.parent.name}: added {added}, skipped {skipped} duplicate(s), cleared file",
                file=sys.stderr,
            )
        except Exception as exc:
            print(f"  [merge] failed to merge {csv_file}: {exc}", file=sys.stderr)

    if total_added or total_skipped:
        print(
            f"  [merge] Total: {total_added} pair(s) added to {master.name}, {total_skipped} skipped",
            file=sys.stderr,
        )


def _health_check_fee_rate(pairs: list[dict], poly, sample_limit: int = 5) -> bool:
    """Verify Polymarket fee-rate endpoint reachability before scanning."""
    checked = 0
    print(f"  [health] checking fee-rate endpoint (sample {sample_limit} pairs)...")
    for pair in pairs:
        if checked >= sample_limit:
            break

        token_ids = [str(t).strip() for t in (pair.get("poly_token_ids") or []) if str(t).strip()]
        if not token_ids:
            y = str(pair.get("polymarket_yes_token_id") or "").strip()
            n = str(pair.get("polymarket_no_token_id") or "").strip()
            if y and n:
                token_ids = [y, n]

        if not token_ids:
            slug = str(pair.get("polymarket_market_slug") or "").strip()
            if slug:
                try:
                    info = poly.resolve_market_outcomes(slug)
                    token_ids = [str(t).strip() for t in (info.get("token_ids") or []) if str(t).strip()]
                except Exception as exc:
                    print(
                        f"  [health] {pair.get('pair_id', '')}: token resolution failed ({exc})",
                        file=sys.stderr,
                    )
                    continue

        if not token_ids:
            continue

        checked += 1
        try:
            rate = _resolve_poly_fee_rate(poly, token_ids)
            print(
                f"  [health] fee-rate ok for {pair.get('pair_id', '')} (rate={rate:.4f})"
            )
            return True
        except Exception as exc:
            print(
                f"  [health] {pair.get('pair_id', '')}: fee-rate check failed ({exc})",
                file=sys.stderr,
            )

    if checked == 0:
        print("  [health] no usable token IDs found to test fee-rate endpoint", file=sys.stderr)
    else:
        print("  [health] fee-rate endpoint check failed on all samples", file=sys.stderr)
    return False


def cmd_scan(args) -> int:
    cfg = load_config(args.config)
    errors = _validate_config(cfg)
    if errors:
        for e in errors:
            print(f"  {Fore.RED}✗{Style.RESET_ALL} {e}", file=sys.stderr)
        return 1

    _compile_fee_fns(cfg)
    cfg["mode"] = _effective_mode(cfg)

    _merge_user_pairs(cfg["pairs_file"], cfg.get("user_pairs_dirs"))
    pairs = load_pairs(cfg["pairs_file"])
    if not pairs:
        print(f"{Fore.YELLOW}No active pairs found in '{cfg['pairs_file']}'{Style.RESET_ALL}")
        return 0

    _print_banner(cfg, len(pairs))
    kalshi, poly = _build_connectors(cfg)
    if args.health_check and not _health_check_fee_rate(pairs, poly):
        print(f"  {Fore.RED}âœ—{Style.RESET_ALL} fee-rate health check failed â€” aborting scan")
        return 1

    # scan-only: no trades placed
    run_scan(pairs, kalshi, poly, cfg, execute=False)
    return 0


def cmd_run(args) -> int:
    cfg = load_config(args.config)
    errors = _validate_config(cfg)
    if errors:
        for e in errors:
            print(f"  {Fore.RED}✗{Style.RESET_ALL} {e}", file=sys.stderr)
        return 1

    _compile_fee_fns(cfg)
    cfg["mode"] = _effective_mode(cfg)

    _merge_user_pairs(cfg["pairs_file"], cfg.get("user_pairs_dirs"))
    pairs = load_pairs(cfg["pairs_file"])
    if not pairs:
        print(f"{Fore.YELLOW}No active pairs found in '{cfg['pairs_file']}'{Style.RESET_ALL}")
        return 0

    _print_banner(cfg, len(pairs))
    kalshi, poly = _build_connectors(cfg)
    if args.health_check and not _health_check_fee_rate(pairs, poly):
        print(f"  {Fore.RED}âœ—{Style.RESET_ALL} fee-rate health check failed â€” aborting run")
        return 1

    run_loop(pairs, kalshi, poly, cfg)
    return 0


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> int:
    colorama_init(autoreset=False)

    parser = argparse.ArgumentParser(
        description="Kalshi × Polymarket cross-market arbitrage bot  (v4)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python main.py validate\n"
            "  python main.py scan\n"
            "  python main.py run\n"
            "  python main.py --config /path/to/config.json run\n"
        ),
    )
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config file (default: config.json). "
             "Falls back to config.example.json if not found.",
    )

    sub = parser.add_subparsers(dest="command", required=True)

    s = sub.add_parser("validate", help="Validate config and pairs file, then exit")
    s.set_defaults(func=cmd_validate)

    s = sub.add_parser(
        "scan",
        help="Run one scan cycle and print opportunities (no trades placed)",
    )
    s.add_argument(
        "--health-check",
        action="store_true",
        help="Verify Polymarket fee-rate endpoint before scanning",
    )
    s.set_defaults(func=cmd_scan)

    s = sub.add_parser(
        "run",
        help="Scan and execute trades continuously until Ctrl+C",
    )
    s.add_argument(
        "--health-check",
        action="store_true",
        help="Verify Polymarket fee-rate endpoint before running",
    )
    s.set_defaults(func=cmd_run)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
