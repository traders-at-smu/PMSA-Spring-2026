#!/usr/bin/env python3
"""Read V2/Pairs_for_Kalshi_and_Polymarket.xlsx and print JSON pairs to stdout.

Usage: python read_manual_pairs.py <path_to_xlsx>
"""

import json
import sys
from pathlib import Path


def _norm(value) -> str:
    return str(value or "").strip()


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: read_manual_pairs.py <path_to_xlsx>"}))
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(json.dumps({"error": f"File not found: {xlsx_path}"}))
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row (row 1)
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [_norm(h).lower() for h in row]

    pairs: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        record = {headers[i]: _norm(v) for i, v in enumerate(row) if i < len(headers)}

        # Skip inactive rows
        active_raw = record.get("active", "").lower()
        if active_raw in {"false", "0", "no", "n"}:
            continue

        # Normalize field names for both schema variants
        kalshi_ticker = record.get("kalshi_market_id") or record.get("kalshi_ticker") or ""
        poly_slug = record.get("poly_slug") or record.get("polymarket_market_slug") or ""
        pair_id = record.get("pair_id") or ""
        title = record.get("title_clean") or record.get("question") or ""
        kalshi_url = record.get("kalshi_url") or (f"https://kalshi.com/markets/{kalshi_ticker}" if kalshi_ticker else "")
        poly_url = record.get("poly_url") or (f"https://polymarket.com/market/{poly_slug}" if poly_slug else "")
        resolution_time = (
            record.get("resolution_time_utc")
            or record.get("expiry_kalshi_utc")
            or record.get("expiry_poly_utc")
            or ""
        )
        category = record.get("category_tag") or record.get("category") or "manual"

        if not kalshi_ticker or not poly_slug:
            continue

        pairs.append({
            "pair_id": pair_id,
            "kalshi_ticker": kalshi_ticker,
            "poly_slug": poly_slug,
            "title": title,
            "kalshi_url": kalshi_url,
            "poly_url": poly_url,
            "resolution_time_utc": resolution_time,
            "category": category,
        })

    wb.close()
    print(json.dumps(pairs))


if __name__ == "__main__":
    main()
